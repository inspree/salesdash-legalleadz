#!/usr/bin/env python3.13
"""
HQ Intake Billing Portal — Flask Web Application
Port 8090

Routes:
  /                  — Admin dashboard (Robert)
  /firm/<firm_id>    — Firm detail page
  /share/<token>     — Client share page (public, token-based)
  /api/refresh       — Re-run data collection
  /api/share/generate/<firm_id> — Generate share token
  /api/invoice/draft/<firm_id>  — Create FreshBooks invoice draft
  /api/invoice/send/<invoice_id> — Send FreshBooks invoice
  /api/leads/<firm_id>          — Get HubSpot leads for a firm
  /api/share/leads/<token>      — Get leads for client share page
  /api/export/all               — Download all firms as Excel
  /api/export/firm/<firm_id>    — Download firm leads as Excel
  /api/export/share/<token>     — Download client share leads as Excel
"""

import json
import io
import os
import subprocess
import sys
import uuid
import hashlib
import time
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from threading import Thread

from functools import wraps
from flask import (Flask, render_template, jsonify, request, redirect,
                   url_for, abort, send_file, Response)

# ── Paths ──
# When deployed to Railway, use relative paths from the app directory
DASHBOARD_DIR = Path(__file__).parent
BASE_DIR = DASHBOARD_DIR
DATA_FILE = DASHBOARD_DIR / "billing_data.json"
SHARE_TOKENS_FILE = DASHBOARD_DIR / "share_tokens.json"
DATA_QUALITY_FILE = DASHBOARD_DIR / "data_quality_log.json"
SALES_SNAPSHOT_TOKENS_FILE = DASHBOARD_DIR / "sales_snapshot_tokens.json"
CREDENTIALS_DIR = DASHBOARD_DIR / "credentials"
FRESHBOOKS_TOKENS_FILE = CREDENTIALS_DIR / "freshbooks_tokens.json"
ENV_FILE = DASHBOARD_DIR / ".env"

# ── Load .env ──
if ENV_FILE.exists():
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line and "PRIVATE_KEY" not in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

HUBSPOT_API_KEY = os.environ.get("HUBSPOT_API_KEY", "")
CONVOSO_TOKEN = "ku0i6z26w423j6p1xxn17fxejapfumv4"
FRESHBOOKS_ACCOUNT_ID = "wgZEN2"

# ── Flask App ──
app = Flask(
    __name__,
    template_folder=str(DASHBOARD_DIR / "templates"),
    static_folder=str(DASHBOARD_DIR / "static"),
    static_url_path="/static",
)
app.jinja_env.globals.update(max=max, min=min, abs=abs, int=int, round=round)

# ── Password Protection ──
DASHBOARD_USER = os.environ.get("DASHBOARD_USER", "admin")
DASHBOARD_PASS = os.environ.get("DASHBOARD_PASS", "")

# Routes that do NOT require auth (client-facing, token-based)
PUBLIC_PREFIXES = (
    "/static/", "/health", "/sales-snapshot/health", "/sales-snapshot/debug",
    "/api/sales-snapshot/leads/", "/api/vendor/", "/api/share/leads/",
    "/share/", "/dashboard/", "/sales-snapshot/",
    "/jre", "/api/jre",
)

def _is_public_route(path):
    """Check if a path should be publicly accessible (no auth)."""
    if any(path.startswith(p) for p in PUBLIC_PREFIXES):
        return True
    # Single-segment paths like /diamond are sales snapshot tokens — public
    parts = path.strip("/").split("/")
    if len(parts) == 1 and parts[0] and parts[0] not in (
        "firm", "api", "static", "favicon.ico"
    ):
        return True
    return False

@app.before_request
def require_auth():
    """Require HTTP Basic Auth for admin routes when DASHBOARD_PASS is set."""
    if not DASHBOARD_PASS:
        return  # No password configured — skip auth
    if _is_public_route(request.path):
        return  # Public route — no auth needed
    auth = request.authorization
    if auth and auth.username == DASHBOARD_USER and auth.password == DASHBOARD_PASS:
        return  # Auth OK
    return Response(
        "Login required", 401,
        {"WWW-Authenticate": 'Basic realm="HQ Intake Dashboard"'})




# ── Data Loading ──
def load_billing_data():
    """Load billing data from JSON file."""
    if DATA_FILE.exists():
        try:
            return json.loads(DATA_FILE.read_text())
        except Exception:
            return {"summary": {}, "firms": {}, "convoso_summary": {}}
    return {"summary": {}, "firms": {}, "convoso_summary": {}}


def load_share_tokens():
    """Load share tokens mapping."""
    if SHARE_TOKENS_FILE.exists():
        try:
            return json.loads(SHARE_TOKENS_FILE.read_text())
        except Exception:
            return {}
    return {}


def save_share_tokens(tokens):
    """Save share tokens mapping."""
    SHARE_TOKENS_FILE.write_text(json.dumps(tokens, indent=2))


def load_data_quality():
    """Load data quality log."""
    if DATA_QUALITY_FILE.exists():
        try:
            return json.loads(DATA_QUALITY_FILE.read_text())
        except Exception:
            return {"issues": [], "last_check": None}
    return {"issues": [], "last_check": None}


def firm_id_from_name(name):
    """Generate a stable firm ID from name."""
    return hashlib.md5(name.encode()).hexdigest()[:12]


def get_firm_by_id(firm_id, data):
    """Find a firm by its ID."""
    for name, firm in data.get("firms", {}).items():
        if firm_id_from_name(name) == firm_id:
            return name, firm
    return None, None


def get_firm_by_token(token):
    """Find a firm by its share token."""
    tokens = load_share_tokens()
    # Token might be the key itself (hash) or the inner UUID
    if token in tokens:
        # The URL token IS the key — need to find the firm_id
        t = tokens[token]
        firm_name = t.get("firm_name", "") if isinstance(t, dict) else t
        return firm_id_from_name(firm_name)
    # Fallback: check inner token field
    for key, t in tokens.items():
        if isinstance(t, dict) and t.get("token") == token:
            firm_name = t.get("firm_name", "")
            return firm_id_from_name(firm_name)
    # Check vendor tokens
    vendor_tokens = load_vendor_tokens()
    if token in vendor_tokens:
        return token  # vendor tokens are self-contained
    return None


def load_vendor_tokens():
    """Load vendor share tokens (filtered by marketing source)."""
    vfile = DASHBOARD_DIR / "vendor_tokens.json"
    if vfile.exists():
        try:
            return json.loads(vfile.read_text())
        except Exception:
            return {}
    return {}


def save_vendor_tokens(tokens):
    """Save vendor share tokens."""
    vfile = DASHBOARD_DIR / "vendor_tokens.json"
    vfile.write_text(json.dumps(tokens, indent=2))


def hubspot_get_leads_by_marketing_source(source_filters):
    """Get leads from HubSpot filtered by marketing_source values."""
    import requests
    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }
    all_leads = []
    for source in source_filters:
        after = None
        while True:
            body = {
                "filterGroups": [{
                    "filters": [{
                        "propertyName": "marketing_source",
                        "operator": "CONTAINS_TOKEN",
                        "value": source
                    }]
                }],
                "properties": [
                    "firstname", "lastname", "email", "phone",
                    "hs_lead_status", "createdate", "marketing_source",
                    "case_type", "notes_last_updated", "rejection_reason",
                    "special_circumstances"
                ],
                "sorts": [{"propertyName": "createdate", "direction": "DESCENDING"}],
                "limit": 100
            }
            if after:
                body["after"] = after
            resp = requests.post(
                "https://api.hubapi.com/crm/v3/objects/contacts/search",
                headers=headers, json=body
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            for c in data.get("results", []):
                props = c.get("properties", {})
                ms = (props.get("marketing_source") or "").lower()
                if any(sf.lower() in ms for sf in source_filters):
                    all_leads.append({
                        "name": f"{props.get('firstname', '')} {props.get('lastname', '')}".strip(),
                        "email": props.get("email", ""),
                        "phone": props.get("phone", ""),
                        "status": props.get("hs_lead_status", "Unknown"),
                        "date": props.get("createdate", "")[:10] if props.get("createdate") else "",
                        "lead_source": props.get("marketing_source", ""),
                        "case_type": props.get("case_type", ""),
                        "notes": props.get("special_circumstances", ""),
                        "rejection_reason": props.get("rejection_reason", "")
                    })
            paging = data.get("paging", {}).get("next", {})
            after = paging.get("after")
            if not after:
                break
    # Deduplicate by phone
    seen = set()
    unique = []
    for l in all_leads:
        key = l.get("phone", "") or l.get("email", "")
        if key and key not in seen:
            seen.add(key)
            unique.append(l)
        elif not key:
            unique.append(l)
    return unique


def hubspot_get_leads_by_deal_name(name_filters):
    """Get leads from HubSpot by searching deals whose name contains the filter terms.
    Used for vendors like Wommster where leads are tracked in deal names, not marketing_source."""
    import requests
    import time
    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }
    all_leads = []
    for name_filter in name_filters:
        after = None
        for _ in range(20):  # max 2000 deals per filter
            body = {
                "filterGroups": [{
                    "filters": [{
                        "propertyName": "dealname",
                        "operator": "CONTAINS_TOKEN",
                        "value": name_filter
                    }]
                }],
                "properties": ["dealname", "dealstage", "createdate"],
                "sorts": [{"propertyName": "createdate", "direction": "DESCENDING"}],
                "limit": 100
            }
            if after:
                body["after"] = str(after)

            for attempt in range(3):
                resp = requests.post(
                    "https://api.hubapi.com/crm/v3/objects/deals/search",
                    headers=headers, json=body
                )
                if resp.status_code == 429:
                    time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                    continue
                break

            if resp.status_code != 200:
                break
            data = resp.json()
            results = data.get("results", [])
            if not results:
                break

            for d in results:
                props = d.get("properties", {})
                dealname = props.get("dealname", "")
                stage_id = props.get("dealstage", "")
                stage_label = DEAL_STAGE_LABELS.get(stage_id, stage_id)

                # Parse contact name and case type from "Name / Case Type - Firm"
                contact_name = dealname.split("/")[0].strip() if "/" in dealname else dealname
                case_type = ""
                if "/" in dealname:
                    rest = dealname.split("/", 1)[1].strip()
                    case_type = rest.split("-")[0].strip() if "-" in rest else rest

                all_leads.append({
                    "name": contact_name,
                    "email": "",
                    "phone": "",
                    "status": stage_label,
                    "date": props.get("createdate", "")[:10] if props.get("createdate") else "",
                    "lead_source": name_filter,
                    "case_type": case_type,
                    "notes": "",
                    "rejection_reason": ""
                })

            after = data.get("paging", {}).get("next", {}).get("after")
            if not after:
                break

    # Deduplicate by name + date
    seen = set()
    unique = []
    for l in all_leads:
        key = f"{l['name']}|{l['date']}"
        if key not in seen:
            seen.add(key)
            unique.append(l)
    return unique


def compute_aging(firms):
    """Compute outstanding aging buckets from invoice data."""
    aging = {"0_30": 0.0, "30_60": 0.0, "60_90": 0.0, "90_plus": 0.0}
    today = datetime.now().date()

    for name, firm in firms.items():
        invoices = firm.get("fb_invoices", [])
        for inv in invoices:
            outstanding = inv.get("outstanding", 0)
            if outstanding <= 0:
                continue
            inv_date_str = inv.get("date", "")
            if not inv_date_str:
                aging["90_plus"] += outstanding
                continue
            try:
                inv_date = datetime.strptime(inv_date_str, "%Y-%m-%d").date()
                days = (today - inv_date).days
                if days <= 30:
                    aging["0_30"] += outstanding
                elif days <= 60:
                    aging["30_60"] += outstanding
                elif days <= 90:
                    aging["60_90"] += outstanding
                else:
                    aging["90_plus"] += outstanding
            except Exception:
                aging["90_plus"] += outstanding

    return aging


def get_firm_health(firm):
    """Compute health status for a firm: green/yellow/red."""
    issues = 0
    if firm.get("fb_total_outstanding", 0) > 0:
        issues += 1
    if firm.get("retainer_alert"):
        issues += 2
    if firm.get("minutes_alert"):
        issues += 2
    rpct = firm.get("retainer_remaining_pct")
    if rpct is not None and rpct <= 30 and rpct > 20:
        issues += 1
    mpct = firm.get("minutes_remaining_pct")
    if mpct is not None and mpct <= 30 and mpct > 20:
        issues += 1

    if issues >= 3:
        return "red"
    elif issues >= 1:
        return "yellow"
    return "green"


# ── FreshBooks Helpers ──
def load_freshbooks_tokens():
    """Load FreshBooks OAuth tokens."""
    if FRESHBOOKS_TOKENS_FILE.exists():
        try:
            return json.loads(FRESHBOOKS_TOKENS_FILE.read_text())
        except Exception:
            return None
    return None


def refresh_freshbooks_token():
    """Refresh FreshBooks OAuth token if expired."""
    import requests
    tokens = load_freshbooks_tokens()
    if not tokens:
        return None

    client_id = os.environ.get("FRESHBOOKS_CLIENT_ID", "")
    client_secret = os.environ.get("FRESHBOOKS_CLIENT_SECRET", "")

    resp = requests.post("https://api.freshbooks.com/auth/oauth/token", json={
        "grant_type": "refresh_token",
        "refresh_token": tokens["refresh_token"],
        "client_id": client_id,
        "client_secret": client_secret,
        "redirect_uri": "https://www.freshbooks.com"
    })

    if resp.status_code == 200:
        new_tokens = resp.json()
        new_tokens["created_at"] = datetime.now(timezone.utc).isoformat()
        FRESHBOOKS_TOKENS_FILE.write_text(json.dumps(new_tokens, indent=2))
        return new_tokens
    return tokens


def fb_headers():
    """Get FreshBooks API headers with valid token."""
    tokens = load_freshbooks_tokens()
    if not tokens:
        return None
    return {
        "Authorization": f"Bearer {tokens['access_token']}",
        "Content-Type": "application/json",
        "Api-Version": "alpha"
    }


def fb_get_client_id(firm_name):
    """Find FreshBooks client ID by firm name."""
    import requests
    headers = fb_headers()
    if not headers:
        return None

    resp = requests.get(
        f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/users/clients",
        headers=headers,
        params={"search[organization_like]": firm_name}
    )
    if resp.status_code == 200:
        clients = resp.json().get("response", {}).get("result", {}).get("clients", [])
        if clients:
            return clients[0].get("id")
    # Try refreshing token
    refresh_freshbooks_token()
    headers = fb_headers()
    if not headers:
        return None
    resp = requests.get(
        f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/users/clients",
        headers=headers,
        params={"search[organization_like]": firm_name}
    )
    if resp.status_code == 200:
        clients = resp.json().get("response", {}).get("result", {}).get("clients", [])
        if clients:
            return clients[0].get("id")
    return None


# ── Deal Stage Mapping ──
DEAL_STAGE_LABELS = {
    "3022527194": "Contacting",
    "qualifiedtobuy": "CB",
    "presentationscheduled": "Rejected",
    "decisionmakerboughtin": "Decision Maker Bought-In",
    "contractsent": "Contacted",
    "closedwon": "Signed",
    "closedlost": "Signed e-Sign",
    "3022527195": "Scheduled Appointment",
    "3022527196": "Signed e-Sign - Commercial",
    "3022527197": "Sent e-Sign",
    "3022527198": "Signed Commercial",
    "3022527199": "HQ Credit",
    "3022527201": "In Call",
    "3022527202": "Dropped",
    "3022527203": "Intake Under Review",
    "3022527204": "DNC",
    "3022527205": "DAIR",
    "3022527206": "Intake Questionnaire Emailed",
    "3022527207": "SC Credit",
    "3022527208": "Temporary Rejection",
    "3022527209": "Cancelled E-Sign",
    "appointmentscheduled": "Rejected",
    "3022527210": "Rejected - Retainer Not Returned",
}


def _get_deal_stages_by_name(firm_name, headers):
    """Search deals by firm name and return {contact_name_lower: {"stage": label, "date": YYYY-MM-DD}}.
    Uses the search API which has its own rate limit bucket. Fast and efficient."""
    import requests
    import time

    name_to_info = {}
    after = 0
    # Deal name format: "Contact Name / Case Type - Firm Name"
    search_token = firm_name.split()[0]

    for _ in range(20):  # max 2000 deals
        body = {
            "filterGroups": [{
                "filters": [{
                    "propertyName": "dealname",
                    "operator": "CONTAINS_TOKEN",
                    "value": search_token
                }]
            }],
            "properties": ["dealname", "dealstage", "createdate"],
            "limit": 100,
        }
        if after:
            body["after"] = str(after)

        for attempt in range(3):
            resp = requests.post(
                "https://api.hubapi.com/crm/v3/objects/deals/search",
                headers=headers,
                json=body
            )
            if resp.status_code == 429:
                time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                continue
            break

        if resp.status_code != 200:
            break

        data = resp.json()
        results = data.get("results", [])
        if not results:
            break

        for d in results:
            props = d.get("properties", {})
            dealname = props.get("dealname", "")
            stage_id = props.get("dealstage", "")
            stage_label = DEAL_STAGE_LABELS.get(stage_id, stage_id)
            deal_date = props.get("createdate", "")[:10] if props.get("createdate") else ""

            # Parse contact name from "Name / Case Type - Firm"
            if "/" in dealname:
                contact_name = dealname.split("/")[0].strip().lower()
                name_to_info[contact_name] = {"stage": stage_label, "date": deal_date}

        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break

    return name_to_info


# ── HubSpot Helpers ──
def hubspot_get_leads_for_firm(firm_name, max_deals=200, since_days=None):
    """Get leads from HubSpot by searching deals for the firm, then fetching contact info.
    This approach uses deals as the source of truth for status and dates.
    max_deals caps the number of deals fetched to prevent timeouts on large firms.
    since_days: if set, only return deals created in the last N days."""
    import requests
    import time

    if not HUBSPOT_API_KEY:
        return []

    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }

    IMPORT_DATE = "2026-02-28"
    # Millisecond timestamp for March 1 — everything before this for Chalik is imported data
    IMPORT_CUTOFF_MS = 1772323200000  # 2026-03-01T00:00:00Z

    # Step 1: Get deals for this firm (paginated, capped at max_deals)
    # Firm name aliases for HubSpot deal name mismatches
    FIRM_SEARCH_ALIASES = {
        "kansas city accident injury attorneys": ["KC", "Accident", "Injury", "Attorneys"],
    }
    alias_words = FIRM_SEARCH_ALIASES.get(firm_name.lower())
    if alias_words:
        search_tokens = alias_words
    else:
        # Use first distinctive word from firm name — deal names only contain
        # abbreviated firm name (e.g., "Schilling & Silvers" not full name)
        SKIP_WORDS = {"the", "law", "office", "of", "a", "and", "llc", "pc", "pllc",
                      "group", "firm", "legal", "services", "associates", "&",
                      "personal", "injury", "car", "accident", "lawyers", "attorney",
                      "attorneys", "at"}
        words = [w.rstrip(".,") for w in firm_name.split()
                 if w.lower().rstrip(".,") not in SKIP_WORDS and len(w.rstrip(".,")) > 1]
        # Use ALL distinctive words with AND logic for precision
        # Previously used first-word-only which caused false positives (e.g., "Daniel"
        # from "Daniel A. Brown" matched every client named Daniel)
        search_tokens = words if words else [firm_name.split()[0]]

    max_pages = max(1, max_deals // 100)
    deals = []
    after = 0
    hubspot_total = 0
    for page_num in range(max_pages):
        name_filters = [
            {"propertyName": "dealname", "operator": "CONTAINS_TOKEN", "value": w}
            for w in search_tokens
        ]
        # Add date filter if since_days is specified
        # Use max(cutoff, IMPORT_CUTOFF) to exclude bulk-imported deals (createdate=Feb 28)
        # whose real dates are months/years old. Only real deals (created Mar 1+) are recent.
        if since_days:
            from datetime import datetime, timedelta
            cutoff = datetime.utcnow() - timedelta(days=since_days)
            cutoff_ms = int(cutoff.timestamp() * 1000)
            # Always use at least March 1 cutoff to skip 15,601 imported deals
            effective_cutoff = max(cutoff_ms, IMPORT_CUTOFF_MS)
            name_filters.append({
                "propertyName": "createdate",
                "operator": "GTE",
                "value": str(effective_cutoff)
            })
        body = {
            "filterGroups": [{"filters": name_filters}],
            "properties": ["dealname", "dealstage", "createdate"],
            "sorts": [{"propertyName": "createdate", "direction": "DESCENDING"}],
            "limit": 100,
        }
        if after:
            body["after"] = str(after)

        for attempt in range(3):
            resp = requests.post(
                "https://api.hubapi.com/crm/v3/objects/deals/search",
                headers=headers, json=body
            )
            if resp.status_code == 429:
                time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                continue
            break

        if resp.status_code != 200:
            break

        data = resp.json()
        if page_num == 0:
            hubspot_total = data.get("total", 0)
        results = data.get("results", [])
        if not results:
            break
        deals.extend(results)
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break

    if not deals:
        return {"leads": [], "hubspot_total": 0}

    # Post-filter: verify deals belong to this firm (check firm portion of dealname)
    FIRM_SEARCH_TOKENS = {
        "KP Injury Law": "KP",
        "The Law Office of Daniel A. Brown": "Brown",
        "Law Office of Shane R. Kadlec": "Kadlec",
        "Law Office of David Kwartler": "Kwartler",
        "AK Law Firm": "AK",
        "Bernard Law Group": "Bernard",
        "Boston Auto Law": "Boston",
        "California Attorney Group": "California",
        "Chalik & Chalik": "Chalik",
        "Fang Accident Lawyers": "Fang",
        "Gibbins Law": "Gibbins",
        "Hollander Law Firm": "Hollander",
        "JRE Injury Law": "JRE",
        "KC Accident Injury Attorneys": "KC",
        "Kansas City Accident Injury Attorneys": "KC",
        "Kronos Law Firm": "Kronos",
        "Larry H. Parker": "Parker",
        "Larry H Parker": "Parker",
        "Levine Law": "Levine",
        "Loncar Lyon Jenkins": "Loncar",
        "Major Law Firm": "Major",
        "The Major Law Firm": "Major",
        "Schilling & Silvers": "Schilling",
        "Schilling & Silvers Personal Injury & Car Accident Lawyers": "Schilling",
        "Shamsi Law Firm": "Shamsi",
        "The Shamsi Law Firm, APC.": "Shamsi",
        "Titan Law Firm": "Titan",
        "Hilley & Solis": "Hilley",
        "Hilley & Solis Law": "Hilley",
        "Pencheff & Fraley": "Pencheff",
        "Klenofsky & Steward": "Klenofsky",
        "Geoff McDonald & Associates": "McDonald",
        "Astrix Law": "Astrix",
        "Jacoby & Meyers": "Jacoby",
        "Edward Law Group": "Edward",
    }
    firm_lower = firm_name.lower()
    ft = FIRM_SEARCH_TOKENS.get(firm_name, "").lower()
    filtered_deals = []
    for d in deals:
        dn = (d.get("properties", {}).get("dealname", "") or "")
        if " - " in dn:
            firm_portion = dn.split(" - ")[-1].strip().lower()
        else:
            firm_portion = dn.lower()
        if firm_lower in firm_portion or (ft and ft in firm_portion):
            filtered_deals.append(d)
    deals = filtered_deals
    hubspot_total = len(deals)

    if not deals:
        return {"leads": [], "hubspot_total": 0}

    # Step 2: Build lead list from deals, fetch contact info in batches
    # First collect deal_id -> {stage, date, contact_name_from_deal}
    deal_contact_ids = {}
    leads = []

    # Get contact associations for deals in batches
    deal_ids = [d["id"] for d in deals]
    deal_props = {}
    for d in deals:
        props = d.get("properties", {})
        deal_props[d["id"]] = {
            "dealname": props.get("dealname", ""),
            "stage": DEAL_STAGE_LABELS.get(props.get("dealstage", ""), props.get("dealstage", "")),
            "date": props.get("createdate", "")[:10] if props.get("createdate") else "",
        }

    # Batch get associations (deal -> contacts) — use larger batches, no sleep
    for batch_start in range(0, len(deal_ids), 100):
        batch = deal_ids[batch_start:batch_start + 100]
        for attempt in range(3):
            assoc_resp = requests.post(
                "https://api.hubapi.com/crm/v4/associations/deals/contacts/batch/read",
                headers=headers,
                json={"inputs": [{"id": did} for did in batch]}
            )
            if assoc_resp.status_code == 429:
                time.sleep(int(assoc_resp.headers.get("Retry-After", 1)) + 1)
                continue
            break
        if assoc_resp.status_code in (200, 207):
            for item in assoc_resp.json().get("results", []):
                did = item.get("from", {}).get("id", "")
                for to in item.get("to", []):
                    cid = to.get("toObjectId", "")
                    if cid and did:
                        deal_contact_ids.setdefault(did, []).append(str(cid))

    # Collect unique contact IDs
    all_cids = set()
    for cids in deal_contact_ids.values():
        all_cids.update(cids)

    # Batch read contacts — use larger batches, no sleep
    contact_data = {}
    cid_list = list(all_cids)
    for batch_start in range(0, len(cid_list), 100):
        batch = cid_list[batch_start:batch_start + 100]
        for attempt in range(3):
            cresp = requests.post(
                "https://api.hubapi.com/crm/v3/objects/contacts/batch/read",
                headers=headers,
                json={
                    "inputs": [{"id": cid} for cid in batch],
                    "properties": [
                        "firstname", "lastname", "email", "phone",
                        "hs_lead_status", "createdate", "lead_source",
                        "notes_last_updated", "rejection_reason",
                        "e_sign_signed_date"
                    ]
                }
            )
            if cresp.status_code == 429:
                time.sleep(int(cresp.headers.get("Retry-After", 1)) + 1)
                continue
            break
        if cresp.status_code in (200, 207):
            for c in cresp.json().get("results", []):
                contact_data[c["id"]] = c.get("properties", {})

    # Step 2b: For deals with no contact association, try to find contacts by name
    unlinked_deal_ids = [did for did in deal_props if did not in deal_contact_ids]
    linked_ratio = len(deal_contact_ids) / max(len(deal_props), 1)
    if linked_ratio < 0.5:
        # Most deals unlinked — skip individual lookups, use deal name as contact name
        unlinked_deal_ids = []
    else:
        unlinked_deal_ids = unlinked_deal_ids[:20]
    if unlinked_deal_ids:
        names_to_search = {}
        for did in unlinked_deal_ids:
            dealname = deal_props[did]["dealname"]
            contact_name = dealname.split("/")[0].strip() if "/" in dealname else dealname
            parts = contact_name.split()
            if len(parts) >= 2:
                names_to_search[did] = {"first": parts[0], "last": " ".join(parts[1:])}

        # Batch search by last name, then match first name
        searched_lastnames = {}
        for did, nm in names_to_search.items():
            ln = nm["last"]
            if ln in searched_lastnames:
                continue
            for attempt in range(3):
                try:
                    sresp = requests.post(
                        "https://api.hubapi.com/crm/v3/objects/contacts/search",
                        headers=headers,
                        json={
                            "filterGroups": [{"filters": [{"propertyName": "lastname", "operator": "EQ", "value": ln}]}],
                            "properties": ["firstname", "lastname", "email", "phone", "lead_source", "rejection_reason"],
                            "limit": 100
                        }
                    )
                    if sresp.status_code == 429:
                        time.sleep(int(sresp.headers.get("Retry-After", 1)) + 1)
                        continue
                    if sresp.status_code == 200:
                        searched_lastnames[ln] = sresp.json().get("results", [])
                    else:
                        searched_lastnames[ln] = []
                    break
                except Exception:
                    searched_lastnames[ln] = []
                    break

        # Match and populate contact_data + deal_contact_ids
        for did, nm in names_to_search.items():
            ln = nm["last"]
            fn_lower = nm["first"].lower()
            for c in searched_lastnames.get(ln, []):
                cp = c.get("properties", {})
                if (cp.get("firstname", "") or "").lower() == fn_lower:
                    cid = c["id"]
                    contact_data[cid] = cp
                    deal_contact_ids.setdefault(did, []).append(cid)
                    break

    # Step 3: Build final lead list from deals + contact info
    seen = set()  # Deduplicate by contact ID
    for did, dp in deal_props.items():
        cids = deal_contact_ids.get(did, [])
        if not cids:
            # No contact linked and no match found — use deal name
            dealname = dp["dealname"]
            if "/" in dealname:
                contact_name = dealname.split("/")[0].strip()
            else:
                contact_name = dealname
            deal_date = dp["date"]
            notes_date = ""
            # For imported deals, we don't have contact_date here (no linked contact)
            # so keep the import date — these will be filtered by the post-filter
            lead_date = deal_date
            leads.append({
                "name": contact_name,
                "email": "",
                "phone": "",
                "status": dp["stage"],
                "date": lead_date,
                "lead_source": "",
                "notes": "",
                "rejection_reason": ""
            })
            continue

        for cid in cids:
            if cid in seen:
                continue
            seen.add(cid)
            props = contact_data.get(cid, {})
            contact_name = f"{props.get('firstname', '')} {props.get('lastname', '')}".strip()
            deal_date = dp["date"]
            esign_date = props.get("e_sign_signed_date", "")[:10] if props.get("e_sign_signed_date") else ""
            stage = dp["stage"]
            is_signed = "sign" in stage.lower() if stage else False

            # Date logic: for signed cases use e-sign date (Forest directive),
            # otherwise use deal createdate (accurate for non-imported deals)
            if is_signed and esign_date:
                lead_date = esign_date
            else:
                lead_date = deal_date

            leads.append({
                "name": contact_name or dp["dealname"].split("/")[0].strip(),
                "email": props.get("email", ""),
                "phone": props.get("phone", ""),
                "status": dp["stage"],
                "date": lead_date,
                "lead_source": props.get("lead_source", ""),
                "notes": props.get("notes_last_updated", ""),
                "rejection_reason": props.get("rejection_reason", "")
            })

    # Post-filter: for signed cases using e_sign_signed_date, ensure the date
    # falls within the requested range (e-sign date may differ from deal createdate)
    if since_days and leads:
        from datetime import datetime, timedelta
        cutoff_date = (datetime.utcnow() - timedelta(days=since_days)).strftime("%Y-%m-%d")
        today_str = datetime.utcnow().strftime("%Y-%m-%d")
        filtered_leads = []
        for lead in leads:
            ld = (lead.get("date") or "")[:10]
            if not ld or (cutoff_date <= ld <= today_str):
                filtered_leads.append(lead)
            # Skip leads whose e-sign date is outside the range
        leads = filtered_leads
        hubspot_total = len(leads)

    return {"leads": leads, "hubspot_total": hubspot_total}


def hubspot_get_leads_from_deals(firm_name, headers):
    """Get leads via deal search when company search fails."""
    import requests

    resp = requests.post(
        "https://api.hubapi.com/crm/v3/objects/deals/search",
        headers=headers,
        json={
            "filterGroups": [{
                "filters": [{
                    "propertyName": "dealname",
                    "operator": "CONTAINS_TOKEN",
                    "value": firm_name.split()[0]
                }]
            }],
            "properties": ["dealname", "dealstage", "createdate", "hs_lastmodifieddate", "amount"],
            "limit": 100
        }
    )

    leads = []
    if resp.status_code == 200:
        deals = resp.json().get("results", [])
        for deal in deals:
            aresp = requests.get(
                f"https://api.hubapi.com/crm/v3/objects/deals/{deal['id']}/associations/contacts",
                headers=headers
            )
            if aresp.status_code == 200:
                contact_ids = [r["id"] for r in aresp.json().get("results", [])]
                for cid in contact_ids[:5]:
                    cresp = requests.get(
                        f"https://api.hubapi.com/crm/v3/objects/contacts/{cid}",
                        headers=headers,
                        params={
                            "properties": "firstname,lastname,email,phone,hs_lead_status,createdate,lastmodifieddate,lead_source,rejection_reason"
                        }
                    )
                    if cresp.status_code in (200, 207):
                        props = cresp.json().get("properties", {})
                        activity_date = props.get("lastmodifieddate", "")[:10] if props.get("lastmodifieddate") else ""
                        create_date = props.get("createdate", "")[:10] if props.get("createdate") else ""
                        lead_date = activity_date or create_date
                        leads.append({
                            "name": f"{props.get('firstname', '')} {props.get('lastname', '')}".strip(),
                            "email": props.get("email", ""),
                            "phone": props.get("phone", ""),
                            "status": props.get("hs_lead_status", "Unknown"),
                            "date": lead_date,
                            "lead_source": props.get("lead_source", ""),
                            "notes": "",
                            "rejection_reason": props.get("rejection_reason", "")
                        })

    return leads


# ── Excel Export Helpers ──
def build_all_firms_excel(data):
    """Build Excel workbook with all firms summary data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "All Firms"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '#,##0.00'
    pct_fmt = '0.0"%"'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "Firm Name", "Status", "Retainers Pre-Paid", "Retainer %",
        "Minutes Pre-Paid", "Minutes Used", "Minutes Remaining", "Minutes %",
        "FB Invoiced", "FB Paid", "FB Outstanding",
        "Convoso Calls (Today)", "Convoso Minutes (Today)",
        "HubSpot Signups", "Sheet Signups", "Source"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    firms = data.get("firms", {})
    convoso = data.get("convoso_summary", {}).get("firms", {})

    row = 2
    for name in sorted(firms.keys(), key=lambda n: firms[n].get("fb_total_outstanding", 0), reverse=True):
        firm = firms[name]
        health = get_firm_health(firm)
        cv = convoso.get(name, {})

        mins_prepaid = firm.get("fb_prepaid_minutes", 0) or 0
        mins_used = firm.get("fb_minutes_used", 0) or 0
        mins_remaining = max(0, mins_prepaid - mins_used)

        values = [
            name,
            health.upper(),
            firm.get("retainers_prepaid", 0),
            firm.get("retainer_remaining_pct"),
            mins_prepaid,
            mins_used,
            mins_remaining,
            firm.get("minutes_remaining_pct"),
            firm.get("fb_total_invoiced", 0),
            firm.get("fb_total_paid", 0),
            firm.get("fb_total_outstanding", 0),
            cv.get("total_calls", 0),
            round(cv.get("total_minutes", 0), 1),
            firm.get("hubspot_signups_since_dec20", 0),
            firm.get("total_signups", 0),
            firm.get("source", "")
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val is not None else "")
            cell.border = thin_border

        # Format currency columns
        for col_idx in [9, 10, 11]:
            ws.cell(row=row, column=col_idx).number_format = currency_fmt

        # Color-code status
        status_cell = ws.cell(row=row, column=2)
        if health == "red":
            status_cell.font = Font(color="FF0000", bold=True)
        elif health == "yellow":
            status_cell.font = Font(color="CC8800", bold=True)
        else:
            status_cell.font = Font(color="228B22", bold=True)

        row += 1

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for r in range(2, row):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Add autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary = data.get("summary", {})
    summary_rows = [
        ("Report Generated", data.get("generated_at", "")[:19]),
        ("Data Range", f"{data.get('data_range', {}).get('start', '')} to {data.get('data_range', {}).get('end', '')}"),
        ("", ""),
        ("Total Firms", summary.get("total_firms", 0)),
        ("FreshBooks Invoiced", summary.get("total_fb_invoiced", 0)),
        ("FreshBooks Paid", summary.get("total_fb_paid", 0)),
        ("FreshBooks Outstanding", summary.get("total_fb_outstanding", 0)),
        ("HubSpot Signups (Dec 20+)", summary.get("hubspot_signups_since_dec20", 0)),
        ("Firms with Alerts", summary.get("firms_with_alerts", 0)),
        ("Convoso Calls Today", summary.get("convoso_total_calls_today", 0)),
        ("Convoso Minutes Today", summary.get("convoso_total_minutes_today", 0)),
    ]
    for r_idx, (label, val) in enumerate(summary_rows, 1):
        ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        cell = ws2.cell(row=r_idx, column=2, value=val)
        if isinstance(val, float) and val > 100:
            cell.number_format = currency_fmt

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_firm_leads_excel(firm_name, leads, include_rejection=True):
    """Build Excel workbook with firm leads."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=f"{firm_name} - Lead Report")
    title_cell.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    headers = ["Name", "Email", "Phone", "Status", "Lead Source", "Date"]
    if include_rejection:
        headers.append("Rejection Reason")
    headers.append("Notes")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r_idx, lead in enumerate(leads, 5):
        values = [
            lead.get("name", ""),
            lead.get("email", ""),
            lead.get("phone", ""),
            lead.get("status", ""),
            lead.get("lead_source", ""),
            lead.get("date", ""),
        ]
        if include_rejection:
            values.append(lead.get("rejection_reason", ""))
        values.append(lead.get("notes", ""))

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val or "")
            cell.border = thin_border

        # Color-code status
        status_cell = ws.cell(row=r_idx, column=4)
        st = (lead.get("status") or "").lower()
        if "signed" in st or "won" in st:
            status_cell.font = Font(color="228B22", bold=True)
        elif "reject" in st or "lost" in st:
            status_cell.font = Font(color="FF0000")
        elif "pending" in st or "new" in st:
            status_cell.font = Font(color="CC8800")

    last_row = 4 + len(leads)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for r in range(5, last_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_client_leads_excel(firm_name, leads):
    """Build client-safe Excel (no billing, no admin data)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Report"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=f"{firm_name} - Lead Report")
    title_cell.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Report Date: {datetime.now().strftime('%B %d, %Y')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    # Summary row
    total = len(leads)
    signed = sum(1 for l in leads if "signed" in (l.get("status") or "").lower() or "won" in (l.get("status") or "").lower())
    pending = sum(1 for l in leads if "pending" in (l.get("status") or "").lower() or "new" in (l.get("status") or "").lower())
    rejected = sum(1 for l in leads if "reject" in (l.get("status") or "").lower() or "lost" in (l.get("status") or "").lower())

    ws.cell(row=3, column=1, value=f"Total: {total}  |  Signed: {signed}  |  Pending: {pending}  |  Rejected: {rejected}")
    ws.cell(row=3, column=1).font = Font(bold=True, size=11)

    headers = ["Name", "Phone", "Email", "Status", "Lead Source", "Date"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r_idx, lead in enumerate(leads, 6):
        values = [
            lead.get("name", ""),
            lead.get("phone", ""),
            lead.get("email", ""),
            lead.get("status", ""),
            lead.get("lead_source", ""),
            lead.get("date", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val or "")
            cell.border = thin_border

        st = (lead.get("status") or "").lower()
        status_cell = ws.cell(row=r_idx, column=4)
        if "signed" in st or "won" in st:
            status_cell.font = Font(color="228B22", bold=True)
        elif "reject" in st or "lost" in st:
            status_cell.font = Font(color="FF0000")

    last_row = 5 + len(leads)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for r in range(6, last_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{last_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Routes ──

@app.route("/")
def dashboard():
    """Admin dashboard — main billing view."""
    data = load_billing_data()
    summary = data.get("summary", {})
    firms = data.get("firms", {})
    convoso = data.get("convoso_summary", {})
    tokens = load_share_tokens()

    # Compute aging
    aging = compute_aging(firms)

    # Build firms list with IDs and health
    firms_list = []
    for name, firm in sorted(firms.items(), key=lambda x: x[1].get("fb_total_outstanding", 0), reverse=True):
        fid = firm_id_from_name(name)
        firm["_id"] = fid
        firm["_name"] = name
        firm["_has_share"] = fid in tokens
        firm["_share_token"] = tokens.get(fid, {}).get("token", "")
        firm["_health"] = get_firm_health(firm)
        firms_list.append(firm)

    # Alerts: below 20% on retainers or minutes
    alert_firms = [f for f in firms_list if f.get("retainer_alert") or f.get("minutes_alert")]

    # Overdue firms (outstanding > 30 days)
    overdue_firms = []
    today = datetime.now().date()
    for f in firms_list:
        if f.get("fb_total_outstanding", 0) > 0:
            invoices = f.get("fb_invoices", [])
            for inv in invoices:
                if inv.get("outstanding", 0) > 0:
                    try:
                        inv_date = datetime.strptime(inv.get("date", ""), "%Y-%m-%d").date()
                        if (today - inv_date).days > 30:
                            overdue_firms.append(f)
                            break
                    except Exception:
                        pass

    # Top outstanding
    top_outstanding = sorted(
        [f for f in firms_list if f.get("fb_total_outstanding", 0) > 0],
        key=lambda x: x.get("fb_total_outstanding", 0), reverse=True
    )[:10]

    # Top signups
    top_signups = sorted(
        [f for f in firms_list if f.get("hubspot_signups_since_dec20", 0) > 0],
        key=lambda x: x.get("hubspot_signups_since_dec20", 0), reverse=True
    )[:10]

    # Data quality issues
    dq = load_data_quality()

    return render_template("dashboard.html",
        summary=summary,
        firms=firms_list,
        alert_firms=alert_firms,
        overdue_firms=overdue_firms,
        top_outstanding=top_outstanding,
        top_signups=top_signups,
        convoso=convoso,
        aging=aging,
        data_quality=dq,
        generated_at=data.get("generated_at", ""),
        data_range=data.get("data_range", {})
    )


@app.route("/firm/<firm_id>")
def firm_detail(firm_id):
    """Firm detail page — full report for one firm."""
    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        abort(404)

    tokens = load_share_tokens()
    share_token = tokens.get(firm_id, {}).get("token", "")
    convoso = data.get("convoso_summary", {}).get("firms", {}).get(name, {})
    health = get_firm_health(firm)

    return render_template("firm_detail.html",
        firm=firm,
        firm_name=name,
        firm_id=firm_id,
        share_token=share_token,
        convoso=convoso,
        hubspot_by_stage=firm.get("hubspot_by_stage", {}),
        invoices=firm.get("fb_invoices", []),
        health=health,
        host=request.host
    )


@app.route("/share/<token>")
def client_share(token):
    """Client share page — token-based, client sees only their leads."""
    # Check vendor tokens first
    vendor_tokens = load_vendor_tokens()
    if token in vendor_tokens:
        vt = vendor_tokens[token]
        return render_template("client_share.html",
            firm_name=vt["vendor_name"],
            firm_id=token,
            token=token,
            firm={"signups": 0}
        )

    firm_id = get_firm_by_token(token)
    if not firm_id:
        abort(404)

    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        abort(404)

    return render_template("client_share.html",
        firm_name=name,
        firm_id=firm_id,
        token=token,
        firm=firm
    )


# ── API Endpoints ──

@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    """Re-run data collection script."""
    script = BASE_DIR / "scripts" / "billing_dashboard_data.py"
    if not script.exists():
        return jsonify({"error": "Data script not found"}), 404

    def run_refresh():
        subprocess.run([sys.executable, str(script)], cwd=str(BASE_DIR))

    thread = Thread(target=run_refresh)
    thread.start()
    return jsonify({"status": "refreshing", "message": "Data refresh started in background"})


@app.route("/api/leads/<firm_id>")
def api_leads(firm_id):
    """Get HubSpot leads for a firm (called via AJAX)."""
    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        return jsonify({"error": "Firm not found"}), 404

    try:
        leads = hubspot_get_leads_for_firm(name)
        return jsonify({"firm": name, "leads": leads, "count": len(leads)})
    except Exception as e:
        return jsonify({"error": str(e), "leads": [], "count": 0}), 500


@app.route("/api/share/leads/<token>")
def api_share_leads(token):
    """Get leads for a share page (client-facing, limited data)."""
    import sys
    print(f"[SHARE] Request for token={token}", flush=True, file=sys.stderr)

    # Check vendor tokens first
    vendor_tokens = load_vendor_tokens()
    if token in vendor_tokens:
        vt = vendor_tokens[token]
        try:
            search_type = vt.get("search_type", "marketing_source")
            if search_type == "deal_name":
                leads = hubspot_get_leads_by_deal_name(vt["filters"])
            else:
                leads = hubspot_get_leads_by_marketing_source(vt["filters"])
            safe_leads = []
            for lead in leads:
                safe_leads.append({
                    "name": lead.get("name", ""),
                    "email": lead.get("email", ""),
                    "phone": lead.get("phone", ""),
                    "status": lead.get("status", "Unknown"),
                    "date": lead.get("date", ""),
                    "lead_source": lead.get("lead_source", ""),
                    "notes": lead.get("notes", ""),
                    "rejection_reason": lead.get("rejection_reason", "")
                })
            return jsonify({"firm": vt["vendor_name"], "leads": safe_leads, "count": len(safe_leads)})
        except Exception as e:
            return jsonify({"error": str(e), "leads": [], "count": 0}), 500

    firm_id = get_firm_by_token(token)
    if not firm_id:
        return jsonify({"error": "Invalid token"}), 404

    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        return jsonify({"error": "Firm not found"}), 404

    try:
        # Accept days parameter from client (30, 60, or 90)
        days = request.args.get('days', 90, type=int)
        if days not in (30, 60, 90):
            days = 90
        # Run with 60-second timeout to prevent Railway request timeout
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(hubspot_get_leads_for_firm, name, 2000, days)
            try:
                result = future.result(timeout=60)
            except FuturesTimeout:
                return jsonify({"error": "Request timed out — too many leads to load. Try a shorter date range.", "leads": [], "count": 0}), 504
        # Handle both old list format and new dict format
        if isinstance(result, dict):
            leads = result.get("leads", [])
            hubspot_total = result.get("hubspot_total", len(leads))
        else:
            leads = result
            hubspot_total = len(leads)
        safe_leads = []
        for lead in leads:
            safe_leads.append({
                "name": lead.get("name", ""),
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "status": lead.get("status", "Unknown"),
                "date": lead.get("date", ""),
                "lead_source": lead.get("lead_source", ""),
                "notes": lead.get("notes", ""),
                "rejection_reason": lead.get("rejection_reason", "")
            })
        return jsonify({"firm": name, "leads": safe_leads, "count": len(safe_leads), "hubspot_total": hubspot_total})
    except Exception as e:
        return jsonify({"error": str(e), "leads": [], "count": 0}), 500


@app.route("/api/share/generate/<firm_id>", methods=["POST"])
def api_generate_share(firm_id):
    """Generate or regenerate a share token for a firm."""
    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        return jsonify({"error": "Firm not found"}), 404

    tokens = load_share_tokens()
    token = uuid.uuid4().hex[:16]
    tokens[firm_id] = {
        "token": token,
        "firm_name": name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    save_share_tokens(tokens)

    share_url = f"{request.scheme}://{request.host}/share/{token}"
    return jsonify({"token": token, "url": share_url, "firm": name})


@app.route("/api/share/generate-all", methods=["POST"])
def api_generate_all_shares():
    """Generate share tokens for all active firms."""
    data = load_billing_data()
    tokens = load_share_tokens()
    generated = 0

    for name, firm in data.get("firms", {}).items():
        fid = firm_id_from_name(name)
        if fid not in tokens:
            if firm.get("fb_total_invoiced", 0) > 0 or firm.get("hubspot_signups_since_dec20", 0) > 0:
                token = uuid.uuid4().hex[:16]
                tokens[fid] = {
                    "token": token,
                    "firm_name": name,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                generated += 1

    save_share_tokens(tokens)
    return jsonify({"status": "ok", "generated": generated, "total": len(tokens)})


# ── Excel Export Endpoints ──

@app.route("/api/export/all")
def api_export_all():
    """Download all firms as Excel."""
    data = load_billing_data()
    output = build_all_firms_excel(data)
    filename = f"HQ_Intake_All_Firms_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/api/export/firm/<firm_id>")
def api_export_firm(firm_id):
    """Download firm leads as Excel."""
    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        abort(404)

    try:
        leads = hubspot_get_leads_for_firm(name)
    except Exception:
        leads = []

    output = build_firm_leads_excel(name, leads)
    safe_name = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')[:30]
    filename = f"{safe_name}_Leads_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/api/export/share/<token>")
def api_export_share(token):
    """Download client-safe leads Excel."""
    firm_id = get_firm_by_token(token)
    if not firm_id:
        abort(404)

    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        abort(404)

    try:
        leads = hubspot_get_leads_for_firm(name)
        safe_leads = [{
            "name": l.get("name", ""),
            "email": l.get("email", ""),
            "phone": l.get("phone", ""),
            "status": l.get("status", ""),
            "lead_source": l.get("lead_source", ""),
            "date": l.get("date", ""),
        } for l in leads]
    except Exception:
        safe_leads = []

    output = build_client_leads_excel(name, safe_leads)
    safe_name = re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')[:30]
    filename = f"{safe_name}_Lead_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ── Invoice Endpoints ──

@app.route("/api/invoice/draft/<firm_id>", methods=["POST"])
def api_invoice_draft(firm_id):
    """Create a FreshBooks invoice draft for a firm."""
    import requests as req

    data = load_billing_data()
    name, firm = get_firm_by_id(firm_id, data)
    if not firm:
        return jsonify({"error": "Firm not found"}), 404

    fb_client_id = fb_get_client_id(name)
    if not fb_client_id:
        return jsonify({"error": f"No FreshBooks client found for '{name}'"}), 404

    lines = []

    minutes_used = firm.get("fb_minutes_used", 0)
    prepaid = firm.get("fb_prepaid_minutes", 0)
    if minutes_used > 0 and prepaid > 0:
        overage = max(0, minutes_used - prepaid)
        if overage > 0:
            lines.append({
                "name": "Minute Overage",
                "description": f"Overage minutes: {overage} mins beyond {prepaid} prepaid",
                "unit_cost": {"amount": "1.50", "code": "USD"},
                "qty": overage
            })

    retainer_pct = firm.get("retainer_remaining_pct")
    if retainer_pct is not None and retainer_pct <= 20:
        signups = firm.get("hubspot_signups_since_dec20", 0) or firm.get("total_signups", 0)
        lines.append({
            "name": "Retainer Renewal -- Signed Cases",
            "description": f"Signed cases for {name}",
            "unit_cost": {"amount": "340.00", "code": "USD"},
            "qty": max(1, signups // 10)
        })

    minutes_pct = firm.get("minutes_remaining_pct")
    if minutes_pct is not None and minutes_pct <= 20:
        lines.append({
            "name": "Prepaid Minutes Renewal -- 2,500 Minutes",
            "description": f"Autodialer minute package for {name}",
            "unit_cost": {"amount": "2694.00", "code": "USD"},
            "qty": 1
        })

    if not lines:
        lines.append({
            "name": "Monthly Service Fee",
            "description": f"HQ Intake services for {name}",
            "unit_cost": {"amount": "2694.00", "code": "USD"},
            "qty": 1
        })

    headers = fb_headers()
    if not headers:
        refresh_freshbooks_token()
        headers = fb_headers()
        if not headers:
            return jsonify({"error": "FreshBooks authentication failed"}), 401

    invoice_data = {
        "invoice": {
            "customerid": fb_client_id,
            "create_date": datetime.now().strftime("%Y-%m-%d"),
            "lines": lines,
            "notes": f"Auto-generated invoice for {name} -- HQ Intake Billing Portal",
            "status": 1
        }
    }

    resp = req.post(
        f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/invoices/invoices",
        headers=headers,
        json=invoice_data
    )

    if resp.status_code in (200, 201):
        inv = resp.json().get("response", {}).get("result", {}).get("invoice", {})
        inv_id = inv.get("id") or inv.get("invoiceid")
        return jsonify({
            "status": "created",
            "invoice_id": inv_id,
            "amount": sum(float(l["unit_cost"]["amount"]) * l["qty"] for l in lines),
            "view_url": f"https://my.freshbooks.com/#/invoice/{inv_id}",
            "firm": name
        })
    else:
        refresh_freshbooks_token()
        headers = fb_headers()
        resp = req.post(
            f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/invoices/invoices",
            headers=headers,
            json=invoice_data
        )
        if resp.status_code in (200, 201):
            inv = resp.json().get("response", {}).get("result", {}).get("invoice", {})
            inv_id = inv.get("id") or inv.get("invoiceid")
            return jsonify({
                "status": "created",
                "invoice_id": inv_id,
                "amount": sum(float(l["unit_cost"]["amount"]) * l["qty"] for l in lines),
                "view_url": f"https://my.freshbooks.com/#/invoice/{inv_id}",
                "firm": name
            })
        return jsonify({"error": f"FreshBooks API error: {resp.status_code}", "detail": resp.text}), 500


@app.route("/api/invoice/send/<invoice_id>", methods=["POST"])
def api_invoice_send(invoice_id):
    """Send a FreshBooks invoice."""
    import requests as req

    headers = fb_headers()
    if not headers:
        refresh_freshbooks_token()
        headers = fb_headers()
        if not headers:
            return jsonify({"error": "FreshBooks authentication failed"}), 401

    resp = req.put(
        f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/invoices/invoices/{invoice_id}",
        headers=headers,
        json={"invoice": {"action_email": True, "status": 2}}
    )

    if resp.status_code == 200:
        return jsonify({"status": "sent", "invoice_id": invoice_id})
    else:
        refresh_freshbooks_token()
        headers = fb_headers()
        resp = req.put(
            f"https://api.freshbooks.com/accounting/account/{FRESHBOOKS_ACCOUNT_ID}/invoices/invoices/{invoice_id}",
            headers=headers,
            json={"invoice": {"action_email": True, "status": 2}}
        )
        if resp.status_code == 200:
            return jsonify({"status": "sent", "invoice_id": invoice_id})
        return jsonify({"error": f"Failed to send: {resp.status_code}", "detail": resp.text}), 500


# ── Sales Snapshot (Multi-Client) ──

def load_sales_snapshot_tokens():
    """Load sales snapshot tokens mapping."""
    if SALES_SNAPSHOT_TOKENS_FILE.exists():
        try:
            return json.loads(SALES_SNAPSHOT_TOKENS_FILE.read_text())
        except Exception:
            return {}
    return {}


def save_sales_snapshot_tokens(tokens):
    """Save sales snapshot tokens mapping."""
    SALES_SNAPSHOT_TOKENS_FILE.write_text(json.dumps(tokens, indent=2))


@app.route("/api/sales-snapshot/generate", methods=["POST"])
def api_generate_sales_snapshot():
    """Generate a sales snapshot token for multiple firms.

    POST JSON body: {"firm_names": ["KP Injury Law", "The Law Office of Daniel A. Brown", ...]}
    OR: {"firm_ids": ["abc123", "def456", ...]}
    """
    payload = request.get_json(force=True) or {}
    data = load_billing_data()
    firms_data = data.get("firms", {})

    firm_names = payload.get("firm_names", [])
    firm_ids = payload.get("firm_ids", [])

    # Resolve firm_ids to firm_names if provided
    if firm_ids and not firm_names:
        for fid in firm_ids:
            name, firm = get_firm_by_id(fid, data)
            if name:
                firm_names.append(name)

    # Validate all firm names exist
    valid_names = []
    for name in firm_names:
        if name in firms_data:
            valid_names.append(name)
        else:
            # Try fuzzy match (case-insensitive contains)
            for fn in firms_data:
                if name.lower() in fn.lower() or fn.lower() in name.lower():
                    valid_names.append(fn)
                    break

    if not valid_names:
        return jsonify({"error": "No valid firms found", "provided": firm_names}), 400

    token = uuid.uuid4().hex[:20]
    tokens = load_sales_snapshot_tokens()
    tokens[token] = {
        "firm_names": valid_names,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    save_sales_snapshot_tokens(tokens)

    share_url = f"{request.scheme}://{request.host}/sales-snapshot/{token}"
    return jsonify({"token": token, "url": share_url, "firms": valid_names})


@app.route("/<token>")
@app.route("/sales-snapshot/<token>")
def sales_snapshot(token):
    """Sales snapshot page -- multi-client overview, token-based, no login required."""
    tokens = load_sales_snapshot_tokens()
    if token not in tokens:
        abort(404)

    token_data = tokens[token]
    firm_names = token_data.get("firm_names", [])

    data = load_billing_data()
    firms_data = data.get("firms", {})

    # Build per-firm context for the template
    firms_context = []
    total_paid_hq_intake = 0.0
    total_ad_spend = 0.0
    ad_spend_map = token_data.get("ad_spend", {})

    # Load QuickBooks cached data for Legal Leadz payments
    qb_data = load_qb_data()

    # ── Month filter: ?month=current|last|2ago ──
    month_filter = request.args.get("month", "")
    month_label = "All Time"
    qb_date_start = None
    qb_date_end = None
    if month_filter in ("current", "last", "2ago"):
        import calendar as _cal
        today = datetime.now()
        if month_filter == "current":
            m, y = today.month, today.year
        elif month_filter == "last":
            m = today.month - 1 if today.month > 1 else 12
            y = today.year if today.month > 1 else today.year - 1
        else:  # 2ago
            dt2 = today.replace(day=1) - timedelta(days=1)
            dt2 = dt2.replace(day=1) - timedelta(days=1)
            m, y = dt2.month, dt2.year
        _, last_day = _cal.monthrange(y, m)
        qb_date_start = f"{y:04d}-{m:02d}-01"
        qb_date_end = f"{y:04d}-{m:02d}-{last_day:02d}"
        month_label = datetime(y, m, 1).strftime("%B %Y")

    def _in_date_range(txn_date_str):
        """Return True if txn_date (YYYY-MM-DD) falls within the selected month."""
        if not qb_date_start:
            return True  # no filter = include all
        if not txn_date_str:
            return False
        return qb_date_start <= txn_date_str[:10] <= qb_date_end

    # Extract QB payments & invoices, optionally filtered by month
    qb_payments_list = (qb_data.get("payments", []) if qb_data else [])
    qb_invoices_list = (qb_data.get("invoices", []) if qb_data else [])
    if qb_date_start:
        qb_payments_list = [p for p in qb_payments_list if _in_date_range(p.get("TxnDate", ""))]
        qb_invoices_list = [i for i in qb_invoices_list if _in_date_range(i.get("TxnDate", ""))]

    # Compute totals from filtered data
    qb_total_payments = sum(float(p.get("TotalAmt", 0)) for p in qb_payments_list)
    qb_payment_count = len(qb_payments_list)

    # "Paid to Legal Leadz" = QB total WITHOUT the 15% management fee
    # Client pays ad_spend * 1.15, so strip the fee: total / 1.15
    total_paid_legal_leadz_actual = round(qb_total_payments / 1.15, 2) if qb_total_payments else 0.0

    # Build per-firm QB payment lookup (fuzzy match on CustomerRef.name)
    # Explicit mapping for QB customer names that don't match firm names
    QB_FIRM_MAP = {
        "cory horne": "kp injury law",
        "james roswold": "kansas city accident injury attorneys",
        "jermey schilling": "schilling & silvers personal injury & car accident lawyers",
        "jeremy schilling": "schilling & silvers personal injury & car accident lawyers",
        "fang law firm pc": "fang accident lawyers",
        "fang law": "fang accident lawyers",
    }

    def _match_firm(customer_name, firm_name):
        """Match QB customer name to firm name, handling middle initials."""
        cn = (customer_name or "").lower().strip()
        fn = (firm_name or "").lower().strip()
        if not cn or not fn:
            return False
        # Check explicit mapping first
        if QB_FIRM_MAP.get(cn) == fn:
            return True
        # Direct substring match
        if fn in cn or cn in fn:
            return True
        # Token-based: strip single-letter tokens (middle initials like "A." or "R.")
        # and check if remaining customer tokens all appear in firm name
        cn_tokens = [t.rstrip(".") for t in cn.split() if len(t.rstrip(".")) > 1]
        fn_tokens = [t.rstrip(".") for t in fn.split() if len(t.rstrip(".")) > 1]
        if cn_tokens and all(t in fn for t in cn_tokens):
            return True
        if fn_tokens and all(t in cn for t in fn_tokens):
            return True
        return False

    firm_qb_payments = {}
    for name in firm_names:
        matched = [p for p in qb_payments_list
                   if _match_firm(p.get("CustomerRef", {}).get("name", ""), name)]
        firm_qb_payments[name] = sum(float(p.get("TotalAmt", 0)) for p in matched)

    for name in firm_names:
        firm = firms_data.get(name, {})
        # FreshBooks total = HQ Intake payments (Forest confirmed)
        paid_hq_fb = firm.get("fb_total_paid", 0.0) or 0.0
        # Ad spend from Google Sheets data (stored in token config)
        firm_ad_spend = ad_spend_map.get(name, 0.0)
        # Per-client Legal Leadz = QB payments for this client / 1.15 (strip 15% fee)
        raw_qb = firm_qb_payments.get(name, 0.0)
        paid_ll = round(raw_qb / 1.15, 2) if raw_qb else 0.0

        total_paid_hq_intake += paid_hq_fb
        total_ad_spend += firm_ad_spend

        firms_context.append({
            "firm_name": name,
            "paid_hq_intake_fb": paid_hq_fb,
            "paid_legal_leadz": paid_ll,
            "fb_invoice_count": firm.get("fb_invoice_count", 0) or 0,
            "total_signups": firm.get("hubspot_signups_since_dec20", 0) or firm.get("total_signups", 0) or 0,
            "ad_spend": firm_ad_spend,
        })

    generated_at = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    return render_template("sales_snapshot.html",
        token=token,
        firms=firms_context,
        firm_names_json=json.dumps(firm_names),
        total_paid_legal_leadz=total_paid_hq_intake,
        total_paid_legal_leadz_actual=total_paid_legal_leadz_actual,
        total_ad_spend=total_ad_spend,
        generated_at=generated_at,
        qb_total_payments=qb_total_payments,
        qb_payment_count=qb_payment_count,
        qb_data_fetched_at=qb_data.get("fetched_at", "") if qb_data else "",
        month_filter=month_filter,
        month_label=month_label,
    )


INJURY_SEVERITY_COLORS = {
    1: "#8BC34A",   # Yellow-green — soft tissue
    2: "#FF9800",   # Orange — moderate (bruising, back issues)
    3: "#FF5722",   # Red-orange — serious (fractures, surgery, TBI)
    4: "#D32F2F",   # Dark red — death, loss of limb, coma
}

# 1-4 scale. Don't underestimate — injuries usually get worse before better.
INJURY_KEYWORDS = {
    4: ["death", "died", "fatal", "deceased", "paralysis", "paralyzed",
        "quadriplegic", "paraplegic", "coma", "brain dead", "catastrophic",
        "amputat", "loss of limb", "lost limb", "lost leg", "lost arm",
        "lost hand", "lost foot", "severed"],
    3: ["surgery", "surgical", "tbi", "traumatic brain", "spinal cord",
        "fracture", "broken", "break", "herniat", "concussion", "torn",
        "ligament", "acl", "mcl", "rotator cuff", "dislocation",
        "multiple injuries", "stitches", "staples", "laceration",
        "internal bleeding", "organ", "ventilator", "icu", "life flight",
        "permanent", "serious", "significant", "hospitalized", "admitted"],
    2: ["bruising", "bruise", "back pain", "back issue", "neck pain",
        "shoulder pain", "knee pain", "whiplash", "sprain", "strain",
        "contusion", "swelling", "limited mobility", "headache",
        "chest pain", "hip pain", "numbness", "tingling", "sciatica",
        "pinched nerve", "bulging disc", "muscle spasm"],
    1: ["sore", "stiff", "tender", "ache", "discomfort", "mild",
        "soft tissue", "minor"],
}

INJURY_PROPERTIES = [
    "what_are_your_injuries",
    "what_are_your_injuries_due_to_this_incident",
    "what_were_your_injuries",
    "what_type_of_injuries_did_you_sustain_must_have_signifcant_injury",
    "please_describe_your_injuries_if_no_breakfracture_stitches_or_surgery_decline",
    "special_circumstances",
    "case_description",
    "did_you_go_to_the_hospital",
    "were_you_transported_to_the_hospital_via_ambulance_ems",
    "have_you_received_treatment_for_these_injuries_erhospital_urgent_care_pcp_chiro",
]


def compute_injury_severity(contact_properties: dict) -> dict:
    """Compute injury severity score (1-4) from HubSpot contact properties.

    Scale: 1=Soft tissue, 2=Moderate (bruising/back), 3=Serious, 4=Death/limb loss/coma
    Default to 2 when no data — injuries usually get worse before better.
    Returns {"score": int, "injuries_text": str, "color": str}.
    """
    # Collect all available injury text
    texts = []
    for prop in INJURY_PROPERTIES:
        val = (contact_properties.get(prop) or "").strip()
        if val and val.lower() not in ("", "none", "n/a", "na", "no"):
            texts.append(val)

    injuries_text = " | ".join(texts)
    combined = injuries_text.lower()

    if not combined:
        # Default to 2 (Moderate) — don't underestimate, injuries get worse
        return {"score": 2, "injuries_text": "(no injury data)", "color": INJURY_SEVERITY_COLORS[2]}

    # Find highest matching severity (check from most severe down)
    best_score = 2  # default to Moderate, not 1 — be conservative
    for score in (4, 3, 2, 1):
        for keyword in INJURY_KEYWORDS[score]:
            if keyword in combined:
                best_score = max(best_score, score)
                break
        if best_score == score and score >= 3:
            break  # found serious+, no need to keep checking

    return {
        "score": best_score,
        "injuries_text": injuries_text[:300],  # cap for tooltip
        "color": INJURY_SEVERITY_COLORS[best_score],
    }



def hubspot_get_signed_deals_by_marketing_source(marketing_sources):
    """Get signed deals filtered by marketing_source values instead of dealname.
    Used for agency tokens like Diamond where the dealname uses the law firm name
    but marketing_source identifies the agency."""
    import requests as req

    if not HUBSPOT_API_KEY:
        return []

    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }

    SIGNED_STAGES = ["closedwon", "closedlost", "3022527196", "3022527198"]

    all_deals = []
    for source in marketing_sources:
        after = 0
        for _ in range(10):
            body = {
                "filterGroups": [{
                    "filters": [
                        {"propertyName": "marketing_source", "operator": "EQ", "value": source},
                        {"propertyName": "dealstage", "operator": "IN", "values": SIGNED_STAGES}
                    ]
                }],
                "properties": ["dealname", "dealstage", "createdate", "marketing_source"],
                "limit": 100,
            }
            if after:
                body["after"] = str(after)

            for attempt in range(3):
                try:
                    resp = req.post(
                        "https://api.hubapi.com/crm/v3/objects/deals/search",
                        headers=headers, json=body, timeout=15
                    )
                except req.exceptions.Timeout:
                    break
                if resp.status_code == 429:
                    time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                    continue
                break

            if resp.status_code != 200:
                break

            data = resp.json()
            results = data.get("results", [])
            for deal in results:
                props = deal.get("properties", {})
                dealname = props.get("dealname", "")
                stage_id = props.get("dealstage", "")
                stage_label = DEAL_STAGE_LABELS.get(stage_id, stage_id)
                date = props.get("createdate", "")[:10] if props.get("createdate") else ""
                contact_name = dealname.split("/")[0].strip() if "/" in dealname else dealname
                case_type = ""
                if "/" in dealname:
                    rest = dealname.split("/", 1)[1].strip()
                    case_type = rest.split("-")[0].strip() if "-" in rest else rest
                all_deals.append({
                    "name": contact_name,
                    "status": stage_label,
                    "date": date,
                    "case_type": case_type,
                    "injury_score": 2,
                    "injury_color": "#FF9800",
                    "injuries_text": "(no injury data)",
                })

            paging = data.get("paging", {}).get("next", {})
            after = paging.get("after")
            if not after:
                break

    return all_deals


def hubspot_get_signed_deals_for_firm(firm_name):
    """Fast version: only fetch deals in signed stages for a firm.
    Much faster than hubspot_get_leads_for_firm() which fetches ALL deals."""
    import requests as req

    if not HUBSPOT_API_KEY:
        return []

    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }

    # Signed deal stages
    SIGNED_STAGES = ["closedwon", "closedlost", "3022527196", "3022527198"]

    # Firm name aliases: when HubSpot deal names use abbreviations or different names
    FIRM_SEARCH_ALIASES = {
        "kansas city accident injury attorneys": ["KC", "Accident", "Injury", "Attorneys"],
    }

    # Check for alias first (case-insensitive)
    alias_words = FIRM_SEARCH_ALIASES.get(firm_name.lower())
    if alias_words:
        search_words = alias_words
    else:
        # Use first distinctive word only — deal names use abbreviated firm names
        # e.g., "Schilling & Silvers" not "Schilling & Silvers Personal Injury..."
        SKIP_WORDS = {"the", "law", "office", "of", "a", "and", "llc", "pc", "pllc",
                      "group", "firm", "legal", "services", "injury", "attorneys", "associates",
                      "personal", "car", "accident", "lawyers", "attorney", "at", "&"}
        search_words = [w.rstrip(".,") for w in firm_name.split()
                        if w.lower().rstrip(".,") not in SKIP_WORDS and len(w.rstrip(".,")) > 1]

        # Use ALL distinctive words with AND logic for precision
        # Previously used first-word-only which caused false positives (e.g., "Daniel"
        # from "Daniel A. Brown" matched every client named Daniel)
        # Aliases handle edge cases like KCAIA where abbreviated names differ
        if not search_words:
            search_words = [firm_name.split()[-1]]

    deals = []
    after = 0

    for _ in range(10):  # max 1000 signed deals
        # Build filters: one CONTAINS_TOKEN per search word + deal stage filter
        name_filters = [
            {"propertyName": "dealname", "operator": "CONTAINS_TOKEN", "value": w}
            for w in search_words
        ]
        body = {
            "filterGroups": [{
                "filters": name_filters + [
                    {
                        "propertyName": "dealstage",
                        "operator": "IN",
                        "values": SIGNED_STAGES
                    }
                ]
            }],
            "properties": ["dealname", "dealstage", "createdate"],
            "limit": 100,
        }
        if after:
            body["after"] = str(after)

        for attempt in range(3):
            try:
                resp = req.post(
                    "https://api.hubapi.com/crm/v3/objects/deals/search",
                    headers=headers, json=body, timeout=15
                )
            except req.exceptions.Timeout:
                break
            if resp.status_code == 429:
                time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                continue
            break
        else:
            break

        if resp.status_code != 200:
            break

        data = resp.json()
        results = data.get("results", [])
        if not results:
            break
        deals.extend(results)
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break

    if not deals:
        return []

    # Post-filter: ensure the firm name actually appears in the FIRM portion of the dealname
    # Dealnames follow "Client Name / Case Type - Firm Name" — only match after " - "
    # This prevents false positives where client names match search tokens
    # (e.g., "Daniel Brown / MVA - Chalik" would falsely match "Daniel A. Brown" firm)
    FIRM_SEARCH_TOKENS = {
        "KP Injury Law": "KP",
        "The Law Office of Daniel A. Brown": "Brown",
        "Law Office of Shane R. Kadlec": "Kadlec",
        "Law Office of David Kwartler": "Kwartler",
        "AK Law Firm": "AK",
        "Bernard Law Group": "Bernard",
        "Boston Auto Law": "Boston",
        "California Attorney Group": "California",
        "Chalik & Chalik": "Chalik",
        "Fang Accident Lawyers": "Fang",
        "Gibbins Law": "Gibbins",
        "Hollander Law Firm": "Hollander",
        "JRE Injury Law": "JRE",
        "KC Accident Injury Attorneys": "KC",
        "Kansas City Accident Injury Attorneys": "KC",
        "Kronos Law Firm": "Kronos",
        "Larry H. Parker": "Parker",
        "Larry H Parker": "Parker",
        "Levine Law": "Levine",
        "Loncar Lyon Jenkins": "Loncar",
        "Major Law Firm": "Major",
        "The Major Law Firm": "Major",
        "Schilling & Silvers": "Schilling",
        "Schilling & Silvers Personal Injury & Car Accident Lawyers": "Schilling",
        "Shamsi Law Firm": "Shamsi",
        "The Shamsi Law Firm, APC.": "Shamsi",
        "Titan Law Firm": "Titan",
        "Hilley & Solis": "Hilley",
        "Hilley & Solis Law": "Hilley",
        "Pencheff & Fraley": "Pencheff",
        "Klenofsky & Steward": "Klenofsky",
        "Geoff McDonald & Associates": "McDonald",
        "Astrix Law": "Astrix",
        "Jacoby & Meyers": "Jacoby",
        "Edward Law Group": "Edward",
    }
    firm_lower = firm_name.lower()
    ft = FIRM_SEARCH_TOKENS.get(firm_name, "").lower()
    filtered_deals = []
    for d in deals:
        dn = (d.get("properties", {}).get("dealname", "") or "")
        if " - " in dn:
            firm_portion = dn.split(" - ")[-1].strip().lower()
        else:
            firm_portion = dn.lower()
        if firm_lower in firm_portion or (ft and ft in firm_portion):
            filtered_deals.append(d)
    deals = filtered_deals

    if not deals:
        return []

    # ── Fetch contact associations & injury properties ──
    deal_ids = [d["id"] for d in deals]
    deal_contact_map = {}  # deal_id -> [contact_id, ...]

    # Batch get associations (deal -> contacts)
    for batch_start in range(0, len(deal_ids), 25):
        batch = deal_ids[batch_start:batch_start + 25]
        try:
            assoc_resp = req.post(
                "https://api.hubapi.com/crm/v4/associations/deals/contacts/batch/read",
                headers=headers,
                json={"inputs": [{"id": did} for did in batch]},
                timeout=15,
            )
            if assoc_resp.status_code in (200, 207):
                for item in assoc_resp.json().get("results", []):
                    did = item.get("from", {}).get("id", "")
                    for to in item.get("to", []):
                        cid = to.get("toObjectId", "")
                        if cid and did:
                            deal_contact_map.setdefault(did, []).append(str(cid))
            elif assoc_resp.status_code == 429:
                time.sleep(int(assoc_resp.headers.get("Retry-After", 2)) + 1)
        except Exception:
            pass
        time.sleep(0.1)

    # Batch read contact injury properties
    all_cids = set()
    for cids in deal_contact_map.values():
        all_cids.update(cids)

    contact_props = {}  # contact_id -> properties dict
    cid_list = list(all_cids)
    for batch_start in range(0, len(cid_list), 50):
        batch = cid_list[batch_start:batch_start + 50]
        try:
            cresp = req.post(
                "https://api.hubapi.com/crm/v3/objects/contacts/batch/read",
                headers=headers,
                json={
                    "inputs": [{"id": cid} for cid in batch],
                    "properties": INJURY_PROPERTIES,
                },
                timeout=15,
            )
            if cresp.status_code in (200, 207):
                for c in cresp.json().get("results", []):
                    contact_props[c["id"]] = c.get("properties", {})
            elif cresp.status_code == 429:
                time.sleep(int(cresp.headers.get("Retry-After", 2)) + 1)
        except Exception:
            pass
        time.sleep(0.1)

    # Build lead list from deal names (parse "Name / Case Type - Firm")
    leads = []
    for d in deals:
        props = d.get("properties", {})
        dealname = props.get("dealname", "")
        stage_id = props.get("dealstage", "")
        stage_label = DEAL_STAGE_LABELS.get(stage_id, stage_id)
        date = props.get("createdate", "")[:10] if props.get("createdate") else ""

        # Parse contact name and case type from dealname
        contact_name = dealname.split("/")[0].strip() if "/" in dealname else dealname
        case_type = ""
        if "/" in dealname:
            rest = dealname.split("/", 1)[1].strip()
            # "Case Type - Firm Name" → extract case type
            case_type = rest.split("-")[0].strip() if "-" in rest else rest

        # Compute injury severity from associated contact
        injury = {"score": 2, "injuries_text": "(no contact data — default Moderate)", "color": INJURY_SEVERITY_COLORS[2]}
        cids = deal_contact_map.get(d["id"], [])
        if cids and cids[0] in contact_props:
            injury = compute_injury_severity(contact_props[cids[0]])

        leads.append({
            "name": contact_name,
            "status": stage_label,
            "date": date,
            "case_type": case_type,
            "injury_score": injury["score"],
            "injury_color": injury["color"],
            "injuries_text": injury["injuries_text"],
        })

    return leads


@app.route("/api/sales-snapshot/leads/<token>/<firm_name>")
def api_sales_snapshot_leads(token, firm_name):
    """Get signed leads for a specific firm within a sales snapshot."""
    tokens = load_sales_snapshot_tokens()
    if token not in tokens:
        return jsonify({"error": "Invalid token"}), 404

    token_data = tokens[token]
    allowed_firms = token_data.get("firm_names", [])

    if firm_name not in allowed_firms:
        return jsonify({"error": "Firm not in snapshot"}), 403

    try:
        marketing_sources = token_data.get("marketing_sources")
        if marketing_sources:
            leads = hubspot_get_signed_deals_by_marketing_source(marketing_sources)
        else:
            leads = hubspot_get_signed_deals_for_firm(firm_name)
        return jsonify({"firm": firm_name, "leads": leads, "count": len(leads)})
    except Exception as e:
        return jsonify({"error": str(e), "leads": [], "count": 0}), 500


@app.route("/sales-snapshot/health")
@app.route("/health")
def sales_snapshot_health():
    return jsonify({"status": "ok", "version": "2026-03-26-v6"})


@app.route("/sales-snapshot/debug")
def sales_snapshot_debug():
    """Debug endpoint to check HubSpot connectivity."""
    import requests as req
    has_key = bool(HUBSPOT_API_KEY)
    key_preview = HUBSPOT_API_KEY[:8] + "..." if has_key else "NOT SET"
    result = {"hubspot_key_set": has_key, "key_preview": key_preview}

    if has_key:
        try:
            resp = req.get(
                "https://api.hubapi.com/crm/v3/objects/deals?limit=1",
                headers={"Authorization": f"Bearer {HUBSPOT_API_KEY}"},
                timeout=10
            )
            result["hubspot_status"] = resp.status_code
            result["hubspot_response"] = resp.text[:200]
        except Exception as e:
            result["hubspot_error"] = str(e)

    return jsonify(result)


# ── QuickBooks OAuth & API Integration ──
import logging

QB_CLIENT_ID = os.environ.get("QUICKBOOKS_CLIENT_ID", "")
QB_CLIENT_SECRET = os.environ.get("QUICKBOOKS_CLIENT_SECRET", "")
QB_REDIRECT_URI = os.environ.get("QUICKBOOKS_REDIRECT_URI", "")
QB_TOKENS_FILE = CREDENTIALS_DIR / "quickbooks_tokens.json"
QB_LOG_FILE = DASHBOARD_DIR / "logs" / "quickbooks_api.log"
QB_SUPPORT_EMAIL = "admin@hqintake.com"

# Set up QuickBooks-specific file logger
(DASHBOARD_DIR / "logs").mkdir(exist_ok=True)
qb_logger = logging.getLogger("quickbooks")
qb_logger.setLevel(logging.DEBUG)
_qb_fh = logging.FileHandler(str(QB_LOG_FILE))
_qb_fh.setFormatter(logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
))
qb_logger.addHandler(_qb_fh)


def load_qb_tokens():
    # Try file first, then fall back to env var (survives Railway redeploys)
    if QB_TOKENS_FILE.exists():
        try:
            return json.loads(QB_TOKENS_FILE.read_text())
        except Exception as e:
            qb_logger.error(f"Failed to load QB tokens from file: {e}")
    env_tokens = os.environ.get("QUICKBOOKS_TOKENS_JSON")
    if env_tokens:
        try:
            tokens = json.loads(env_tokens)
            # Restore to file for faster subsequent loads
            CREDENTIALS_DIR.mkdir(exist_ok=True)
            QB_TOKENS_FILE.write_text(json.dumps(tokens, indent=2))
            qb_logger.info("Restored QB tokens from env var")
            return tokens
        except Exception as e:
            qb_logger.error(f"Failed to load QB tokens from env: {e}")
    return None


def save_qb_tokens(tokens):
    CREDENTIALS_DIR.mkdir(exist_ok=True)
    QB_TOKENS_FILE.write_text(json.dumps(tokens, indent=2))
    os.environ["QUICKBOOKS_TOKENS_JSON"] = json.dumps(tokens)
    # Persist to Railway env var so tokens survive redeploys
    _persist_qb_tokens_to_railway(tokens)
    qb_logger.info("QuickBooks tokens saved successfully")


def _persist_qb_tokens_to_railway(tokens):
    """Save QB tokens as a Railway environment variable via the Railway API."""
    import requests as req
    rt = os.environ.get("RAILWAY_TOKEN")
    pid = os.environ.get("RAILWAY_PROJECT_ID")
    sid = os.environ.get("RAILWAY_SERVICE_ID")
    eid = os.environ.get("RAILWAY_ENV_ID")
    if not all([rt, pid, sid, eid]):
        qb_logger.warning("Railway API vars not set, skipping token persistence")
        return
    try:
        tokens_json = json.dumps(tokens).replace("\\", "\\\\").replace('"', '\\"')
        query = (
            'mutation { variableUpsert(input: { '
            f'projectId: "{pid}", environmentId: "{eid}", '
            f'serviceId: "{sid}", name: "QUICKBOOKS_TOKENS_JSON", '
            f'value: "{tokens_json}" '
            '}) }'
        )
        resp = req.post("https://backboard.railway.com/graphql/v2",
            headers={"Authorization": f"Bearer {rt}", "Content-Type": "application/json"},
            json={"query": query}, timeout=10)
        if resp.status_code == 200 and resp.json().get("data", {}).get("variableUpsert"):
            qb_logger.info("QB tokens persisted to Railway env var")
        else:
            qb_logger.error(f"Railway token persist failed: {resp.text[:200]}")
    except Exception as e:
        qb_logger.error(f"Railway token persist error: {e}")


def _log_qb_response(resp, context=""):
    """Log intuit_tid and response details from any QuickBooks API response."""
    intuit_tid = resp.headers.get("intuit_tid", "N/A")
    qb_logger.info(
        f"[{context}] status={resp.status_code} intuit_tid={intuit_tid}"
    )
    if resp.status_code >= 400:
        qb_logger.error(
            f"[{context}] ERROR status={resp.status_code} intuit_tid={intuit_tid} "
            f"body={resp.text[:1000]}"
        )
    return intuit_tid


def _qb_error_response(status_code, message, intuit_tid="N/A", context=""):
    """Build a standardized error response for QuickBooks API errors."""
    error_detail = {
        "error": message,
        "intuit_tid": intuit_tid,
        "support": QB_SUPPORT_EMAIL,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    qb_logger.error(f"[{context}] Returning error to client: {error_detail}")
    return jsonify(error_detail), status_code


def refresh_qb_token():
    tokens = load_qb_tokens()
    if not tokens or not tokens.get("refresh_token"):
        qb_logger.warning("Token refresh requested but no refresh_token available")
        return None
    import requests as req
    import base64
    auth = base64.b64encode(f"{QB_CLIENT_ID}:{QB_CLIENT_SECRET}".encode()).decode()
    try:
        resp = req.post("https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
            headers={"Authorization": f"Basic {auth}", "Content-Type": "application/x-www-form-urlencoded"},
            data={"grant_type": "refresh_token", "refresh_token": tokens["refresh_token"]},
            timeout=15)
        intuit_tid = _log_qb_response(resp, context="token_refresh")
        if resp.status_code == 200:
            new_tokens = resp.json()
            new_tokens["realm_id"] = tokens.get("realm_id", "")
            new_tokens["updated_at"] = datetime.now().isoformat()
            save_qb_tokens(new_tokens)
            return new_tokens
        qb_logger.error(
            f"Token refresh failed: status={resp.status_code} intuit_tid={intuit_tid} "
            f"body={resp.text[:500]}"
        )
    except Exception as e:
        qb_logger.error(f"Token refresh exception: {e}")
    return None


def qb_api_request(method, url, **kwargs):
    """Make a QuickBooks API request with automatic token refresh and error handling.

    Returns (response, intuit_tid) on success, or raises an exception with logged details.
    Handles: 401 (auto-refresh), 403, 404, syntax/validation errors.
    """
    import requests as req
    tokens = load_qb_tokens()
    if not tokens or not tokens.get("access_token"):
        qb_logger.error("No access token available for API request")
        return None, None

    headers = kwargs.pop("headers", {})
    headers.setdefault("Authorization", f"Bearer {tokens['access_token']}")
    headers.setdefault("Accept", "application/json")
    headers.setdefault("Content-Type", "application/json")

    context = kwargs.pop("context", url)

    try:
        resp = req.request(method, url, headers=headers, timeout=kwargs.pop("timeout", 30), **kwargs)
        intuit_tid = _log_qb_response(resp, context=context)

        # 401 — token expired, try refresh once
        if resp.status_code == 401:
            qb_logger.info(f"[{context}] 401 received, attempting token refresh...")
            new_tokens = refresh_qb_token()
            if new_tokens:
                headers["Authorization"] = f"Bearer {new_tokens['access_token']}"
                resp = req.request(method, url, headers=headers, timeout=30, **kwargs)
                intuit_tid = _log_qb_response(resp, context=f"{context}_retry")
            else:
                qb_logger.error(f"[{context}] Token refresh failed, cannot retry")

        # Log specific error categories
        if resp.status_code == 403:
            qb_logger.error(
                f"[{context}] 403 Forbidden — check app permissions. intuit_tid={intuit_tid}"
            )
        elif resp.status_code == 404:
            qb_logger.error(
                f"[{context}] 404 Not Found — resource does not exist. intuit_tid={intuit_tid}"
            )
        elif resp.status_code == 400:
            # Syntax/validation errors from QuickBooks
            try:
                err_body = resp.json()
                fault = err_body.get("Fault", {})
                errors = fault.get("Error", [])
                for err in errors:
                    qb_logger.error(
                        f"[{context}] Validation error: code={err.get('code')} "
                        f"element={err.get('element', 'N/A')} "
                        f"message={err.get('Message', '')} "
                        f"detail={err.get('Detail', '')} intuit_tid={intuit_tid}"
                    )
            except Exception:
                qb_logger.error(
                    f"[{context}] 400 Bad Request: {resp.text[:500]} intuit_tid={intuit_tid}"
                )

        return resp, intuit_tid

    except req.exceptions.Timeout:
        qb_logger.error(f"[{context}] Request timed out")
        return None, None
    except req.exceptions.ConnectionError as e:
        qb_logger.error(f"[{context}] Connection error: {e}")
        return None, None
    except Exception as e:
        qb_logger.error(f"[{context}] Unexpected error: {e}")
        return None, None


@app.route("/quickbooks/connect")
def qb_connect():
    """Start QuickBooks OAuth flow."""
    if not QB_CLIENT_ID:
        qb_logger.error("OAuth connect attempted but QUICKBOOKS_CLIENT_ID not set")
        return (f"QuickBooks Client ID not configured. "
                f"Contact {QB_SUPPORT_EMAIL} for assistance."), 500
    import urllib.parse
    params = urllib.parse.urlencode({
        "client_id": QB_CLIENT_ID,
        "response_type": "code",
        "scope": "com.intuit.quickbooks.accounting",
        "redirect_uri": QB_REDIRECT_URI,
        "state": hashlib.sha256(QB_CLIENT_ID.encode()).hexdigest()[:16],
    })
    qb_logger.info("OAuth flow initiated")
    return redirect(f"https://appcenter.intuit.com/connect/oauth2?{params}")


@app.route("/quickbooks/callback", methods=["GET", "POST"])
def qb_callback():
    """Handle QuickBooks OAuth callback (GET) and webhook events (POST)."""
    import requests as req
    import base64

    # POST = Intuit webhook notification
    if request.method == "POST":
        import hmac as hmac_mod
        verifier = os.environ.get("QUICKBOOKS_VERIFIER_TOKEN", "")
        signature = request.headers.get("intuit-signature", "")
        payload = request.get_data()
        qb_logger.info(f"Webhook received: {len(payload)} bytes")
        if verifier and signature:
            expected = base64.b64encode(
                hmac_mod.new(verifier.encode(), payload, hashlib.sha256).digest()
            ).decode()
            if not hmac_mod.compare_digest(expected, signature):
                qb_logger.warning("Webhook signature verification FAILED")
                return "Invalid signature", 401
        qb_logger.info(f"Webhook processed OK: {payload[:200]}")
        return "OK", 200

    # GET = OAuth callback
    code = request.args.get("code")
    realm_id = request.args.get("realmId")
    state = request.args.get("state")
    error = request.args.get("error")

    if error:
        qb_logger.error(f"OAuth denied by user: {error}")
        return (f"QuickBooks authorization denied: {error}. "
                f"Contact {QB_SUPPORT_EMAIL} for assistance."), 400
    if not code:
        qb_logger.error("OAuth callback missing authorization code")
        return (f"Missing authorization code. "
                f"Contact {QB_SUPPORT_EMAIL} for assistance."), 400

    auth = base64.b64encode(f"{QB_CLIENT_ID}:{QB_CLIENT_SECRET}".encode()).decode()
    try:
        resp = req.post("https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
            headers={"Authorization": f"Basic {auth}", "Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": QB_REDIRECT_URI,
            },
            timeout=15)
        intuit_tid = _log_qb_response(resp, context="oauth_token_exchange")
    except Exception as e:
        qb_logger.error(f"OAuth token exchange exception: {e}")
        return (f"Token exchange failed. Contact {QB_SUPPORT_EMAIL} for assistance. "
                f"Error: {e}"), 500

    if resp.status_code != 200:
        return (f"Token exchange failed (status {resp.status_code}, intuit_tid={intuit_tid}). "
                f"Contact {QB_SUPPORT_EMAIL} for assistance."), 500

    tokens = resp.json()
    tokens["realm_id"] = realm_id or ""
    tokens["connected_at"] = datetime.now().isoformat()
    save_qb_tokens(tokens)

    qb_logger.info(f"QuickBooks connected: realm_id={realm_id} intuit_tid={intuit_tid}")
    return f"""<html><body style="font-family:sans-serif;text-align:center;padding:60px;">
    <h1 style="color:#059669;">QuickBooks Connected!</h1>
    <p>Company ID: {realm_id}</p>
    <p>Access token and refresh token saved. The dashboard will now pull Legal Leadz payment data automatically.</p>
    <p><a href="/">Back to Dashboard</a></p>
    <p style="color:#666;font-size:0.85em;margin-top:40px;">
        Need help? Contact <a href="mailto:{QB_SUPPORT_EMAIL}">{QB_SUPPORT_EMAIL}</a>
    </p>
    </body></html>"""


@app.route("/quickbooks/export-tokens")
def qb_export_tokens():
    """Export QB tokens as JSON (admin only, protected by secret)."""
    secret = request.args.get("secret", "")
    admin_secret = os.environ.get("ADMIN_SECRET", "")
    if not admin_secret or secret != admin_secret:
        abort(403)
    tokens = load_qb_tokens()
    if not tokens:
        return jsonify({"error": "No tokens"}), 404
    return jsonify(tokens)


@app.route("/quickbooks/status")
def qb_status():
    """Check QuickBooks connection status."""
    tokens = load_qb_tokens()
    if not tokens:
        return jsonify({
            "connected": False,
            "connect_url": "/quickbooks/connect",
            "support": QB_SUPPORT_EMAIL,
        })
    return jsonify({
        "connected": True,
        "realm_id": tokens.get("realm_id", ""),
        "connected_at": tokens.get("connected_at", ""),
        "has_refresh_token": bool(tokens.get("refresh_token")),
        "support": QB_SUPPORT_EMAIL,
    })


# ── QuickBooks Data Routes ──
QB_DATA_FILE = DASHBOARD_DIR / "quickbooks_data.json"


def _qb_query(entity, max_results=1000):
    """Run a QuickBooks query for a given entity type (Invoice, Payment, etc.).
    Returns parsed JSON response or None on failure.
    """
    tokens = load_qb_tokens()
    if not tokens or not tokens.get("realm_id"):
        return None, "QuickBooks not connected"

    realm_id = tokens["realm_id"]
    import urllib.parse
    query = f"SELECT * FROM {entity} MAXRESULTS {max_results}"
    url = (
        f"https://quickbooks.api.intuit.com/v3/company/{realm_id}/query"
        f"?query={urllib.parse.quote(query)}"
    )
    resp, intuit_tid = qb_api_request("GET", url, context=f"query_{entity}")
    if resp is None:
        return None, f"QuickBooks API request failed (no response)"
    if resp.status_code != 200:
        return None, (
            f"QuickBooks API error: status={resp.status_code} "
            f"intuit_tid={intuit_tid}"
        )
    try:
        return resp.json(), None
    except Exception as e:
        return None, f"Failed to parse QB response: {e}"


@app.route("/quickbooks/invoices")
def qb_invoices():
    """Query QuickBooks for all invoices and return as JSON."""
    data, error = _qb_query("Invoice")
    if error:
        return _qb_error_response(502, error, context="qb_invoices")
    invoices = (
        data.get("QueryResponse", {}).get("Invoice", [])
        if data else []
    )
    return jsonify({
        "count": len(invoices),
        "invoices": invoices,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    })


@app.route("/quickbooks/payments")
def qb_payments():
    """Query QuickBooks for all payments received and return as JSON."""
    data, error = _qb_query("Payment")
    if error:
        return _qb_error_response(502, error, context="qb_payments")
    payments = (
        data.get("QueryResponse", {}).get("Payment", [])
        if data else []
    )
    return jsonify({
        "count": len(payments),
        "payments": payments,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    })


# Clients to include in dashboard QB data (empty = include ALL for now)
# Will be set to specific QB customer names once we identify the correct ones
QB_DASHBOARD_CLIENTS = []


def _qb_matches_client(record, client_names):
    """Check if a QB invoice/payment CustomerRef matches any dashboard client."""
    if not client_names:
        return True
    cust_name = (record.get("CustomerRef", {}).get("name", "") or "").lower()
    for name in client_names:
        if name.lower() in cust_name or cust_name in name.lower():
            return True
    return False


@app.route("/api/quickbooks/refresh-data", methods=["POST"])
def qb_refresh_data():
    """Pull QB invoice + payment data filtered to dashboard clients only."""
    tokens = load_qb_tokens()
    if not tokens or not tokens.get("realm_id"):
        return _qb_error_response(400, "QuickBooks not connected", context="qb_refresh_data")

    # Fetch invoices
    inv_data, inv_err = _qb_query("Invoice")
    all_invoices = []
    if inv_data and not inv_err:
        all_invoices = inv_data.get("QueryResponse", {}).get("Invoice", [])

    # Fetch payments
    pay_data, pay_err = _qb_query("Payment")
    all_payments = []
    if pay_data and not pay_err:
        all_payments = pay_data.get("QueryResponse", {}).get("Payment", [])

    # Filter to dashboard clients only
    invoices = [inv for inv in all_invoices if _qb_matches_client(inv, QB_DASHBOARD_CLIENTS)]
    payments = [p for p in all_payments if _qb_matches_client(p, QB_DASHBOARD_CLIENTS)]

    # Compute summary metrics
    total_invoiced = sum(float(inv.get("TotalAmt", 0)) for inv in invoices)
    total_paid = sum(float(p.get("TotalAmt", 0)) for p in payments)
    total_balance = sum(float(inv.get("Balance", 0)) for inv in invoices)
    overdue_invoices = [
        inv for inv in invoices
        if float(inv.get("Balance", 0)) > 0
        and inv.get("DueDate")
        and inv["DueDate"] < datetime.now().strftime("%Y-%m-%d")
    ]

    result = {
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "realm_id": tokens.get("realm_id", ""),
        "filtered_clients": QB_DASHBOARD_CLIENTS,
        "unfiltered_totals": {
            "invoice_count": len(all_invoices),
            "payment_count": len(all_payments),
        },
        "summary": {
            "total_invoiced": round(total_invoiced, 2),
            "total_payments_received": round(total_paid, 2),
            "total_outstanding_balance": round(total_balance, 2),
            "invoice_count": len(invoices),
            "payment_count": len(payments),
            "overdue_count": len(overdue_invoices),
        },
        "invoices": invoices,
        "payments": payments,
    }

    # Save to file
    try:
        QB_DATA_FILE.write_text(json.dumps(result, indent=2))
        qb_logger.info(
            f"QuickBooks data refreshed: {len(invoices)} invoices, "
            f"{len(payments)} payments, total_paid=${total_paid:.2f}"
        )
    except Exception as e:
        qb_logger.error(f"Failed to save quickbooks_data.json: {e}")
        return _qb_error_response(500, f"Data fetched but failed to save: {e}",
                                  context="qb_refresh_data")

    errors = []
    if inv_err:
        errors.append(f"Invoices: {inv_err}")
    if pay_err:
        errors.append(f"Payments: {pay_err}")

    return jsonify({
        "status": "ok" if not errors else "partial",
        "summary": result["summary"],
        "errors": errors if errors else None,
        "fetched_at": result["fetched_at"],
    })


def load_qb_data():
    """Load cached QuickBooks data from quickbooks_data.json."""
    if QB_DATA_FILE.exists():
        try:
            return json.loads(QB_DATA_FILE.read_text())
        except Exception:
            pass
    return None


@app.route("/api/quickbooks/debug-customers")
def api_qb_debug_customers():
    """Debug: show QB payment customer names and matching results."""
    qb_data = load_qb_data()
    if not qb_data:
        return jsonify({"error": "No QB data cached"})

    payments = qb_data.get("payments", [])
    data = load_billing_data()
    firms_data = data.get("firms", {})
    all_firm_names = sorted(firms_data.keys())

    # Show each payment's customer and what it matches
    results = []
    for p in payments:
        cref = p.get("CustomerRef", {})
        cname = cref.get("name", "?")
        amt = p.get("TotalAmt", 0)
        # Find which firm name(s) it matches
        matches = []
        for fn in all_firm_names:
            if _match_firm_global(cname, fn):
                matches.append(fn)
        results.append({
            "customer": cname,
            "amount": amt,
            "matched_firms": matches
        })

    return jsonify({
        "total_payments": len(payments),
        "total_firms": len(all_firm_names),
        "payments": results
    })


# Global version of _match_firm for the debug endpoint
def _match_firm_global(customer_name, firm_name):
    """Match QB customer name to firm name."""
    QB_FIRM_MAP = {"cory horne": "kp injury law"}
    cn = (customer_name or "").lower().strip()
    fn = (firm_name or "").lower().strip()
    if not cn or not fn:
        return False
    if QB_FIRM_MAP.get(cn) == fn:
        return True
    if fn in cn or cn in fn:
        return True
    cn_tokens = [t.rstrip(".") for t in cn.split() if len(t.rstrip(".")) > 1]
    fn_tokens = [t.rstrip(".") for t in fn.split() if len(t.rstrip(".")) > 1]
    if cn_tokens and all(t in fn for t in cn_tokens):
        return True
    if fn_tokens and all(t in cn for t in fn_tokens):
        return True
    return False


# ── Vendor Dashboard (Shamsi-style, full deal view) ──

VENDOR_DEAL_PROPERTIES = [
    "dealname", "dealstage", "createdate", "closedate",
    "case_type", "marketing_source", "special_circumstances",
    "notes_last_updated",
]
VENDOR_CONTACT_PROPERTIES = [
    "firstname", "lastname", "email", "phone",
    "case_type", "marketing_source", "special_circumstances",
    "hs_lead_status", "notes_last_updated",
]
VENDOR_DEAL_STAGE_MAP = {}  # Populated at startup from HubSpot pipeline API
_stage_map_loaded = False


def _load_stage_map_from_hubspot():
    """Fetch deal stage labels directly from HubSpot pipeline API so labels match exactly."""
    global VENDOR_DEAL_STAGE_MAP, _stage_map_loaded
    if _stage_map_loaded and VENDOR_DEAL_STAGE_MAP:
        return
    try:
        import requests as _req
        resp = _req.get(
            "https://api.hubapi.com/crm/v3/pipelines/deals",
            headers={"Authorization": f"Bearer {HUBSPOT_API_KEY}"},
            timeout=10,
        )
        if resp.status_code == 200:
            for pipeline in resp.json().get("results", []):
                for stage in pipeline.get("stages", []):
                    VENDOR_DEAL_STAGE_MAP[stage["id"]] = stage["label"]
            _stage_map_loaded = True
            print(f"[Stage Map] Loaded {len(VENDOR_DEAL_STAGE_MAP)} stages from HubSpot")
    except Exception as e:
        print(f"[Stage Map] Error: {e} — using fallback")
        VENDOR_DEAL_STAGE_MAP.update({
            "3022527194": "Contacting", "qualifiedtobuy": "CB",
            "presentationscheduled": "RJCTD", "contractsent": "CNTCT",
            "closedwon": "SISIGN", "closedlost": "Signed e-Sign",
            "3022527196": "Signed e-Sign - Commercial",
            "appointmentscheduled": "Rejected",
        })


def hubspot_get_vendor_deals(firm_names, month_offset=0, max_deals=500):
    """Fetch deals for firm(s) with full contact info — Shamsi-style dashboard data.
    Returns list of deal dicts matching the vendor dashboard format."""
    import requests as req

    _load_stage_map_from_hubspot()

    if not HUBSPOT_API_KEY:
        return [], {"total": 0, "by_source": {}, "by_stage": {}}, "N/A", []

    headers = {
        "Authorization": f"Bearer {HUBSPOT_API_KEY}",
        "Content-Type": "application/json"
    }

    # Calculate date range for month filter
    from datetime import datetime, timedelta
    now = datetime.utcnow()
    target_month = now.month + month_offset
    target_year = now.year
    while target_month <= 0:
        target_month += 12
        target_year -= 1
    while target_month > 12:
        target_month -= 12
        target_year += 1

    # Start of target month
    month_start = datetime(target_year, target_month, 1)
    # Start of next month
    if target_month == 12:
        month_end = datetime(target_year + 1, 1, 1)
    else:
        month_end = datetime(target_year, target_month + 1, 1)

    month_label = month_start.strftime("%B %Y")
    start_ms = int(month_start.timestamp() * 1000)
    end_ms = int(month_end.timestamp() * 1000)

    # Firm name aliases
    FIRM_SEARCH_ALIASES = {
        "kansas city accident injury attorneys": ["KC", "Accident", "Injury", "Attorneys"],
    }
    SKIP_WORDS = {"the", "law", "office", "of", "a", "and", "llc", "pc", "pllc",
                  "group", "firm", "legal", "services", "injury", "attorneys", "associates",
                  "personal", "car", "accident", "lawyers", "attorney", "at", "&"}

    # Build search words from ALL firm names (OR logic across firms)
    filter_groups = []
    for firm_name in firm_names:
        alias = FIRM_SEARCH_ALIASES.get(firm_name.lower())
        if alias:
            words = alias
        else:
            words = [w.rstrip(".,") for w in firm_name.split()
                     if w.lower().rstrip(".,") not in SKIP_WORDS and len(w.rstrip(".,")) > 1]
            if not words:
                words = [firm_name.split()[-1]]

        name_filters = [
            {"propertyName": "dealname", "operator": "CONTAINS_TOKEN", "value": w}
            for w in words
        ]
        name_filters.append({
            "propertyName": "createdate", "operator": "GTE", "value": str(start_ms)
        })
        name_filters.append({
            "propertyName": "createdate", "operator": "LT", "value": str(end_ms)
        })
        filter_groups.append({"filters": name_filters})

    # Fetch deals
    deals = []
    after = 0
    max_pages = max(1, max_deals // 100)

    for _ in range(max_pages):
        body = {
            "filterGroups": filter_groups,
            "properties": VENDOR_DEAL_PROPERTIES,
            "sorts": [{"propertyName": "createdate", "direction": "DESCENDING"}],
            "limit": 100,
        }
        if after:
            body["after"] = str(after)

        for attempt in range(3):
            try:
                resp = req.post(
                    "https://api.hubapi.com/crm/v3/objects/deals/search",
                    headers=headers, json=body, timeout=20
                )
            except req.exceptions.Timeout:
                break
            if resp.status_code == 429:
                time.sleep(int(resp.headers.get("Retry-After", 2)) + 1)
                continue
            break
        else:
            break

        if resp.status_code != 200:
            break

        data = resp.json()
        results = data.get("results", [])
        if not results:
            break
        deals.extend(results)
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break

    if not deals:
        return [], {"total": 0, "by_source": {}, "by_stage": {}}, month_label, []

    # Batch fetch contact associations
    deal_ids = [d["id"] for d in deals]
    deal_contact_map = {}

    import logging
    _vlog = logging.getLogger('vendor_dash')
    for batch_start in range(0, len(deal_ids), 25):
        batch = deal_ids[batch_start:batch_start + 25]
        batch_num = batch_start // 25 + 1
        for attempt in range(3):
            try:
                assoc_resp = req.post(
                    "https://api.hubapi.com/crm/v4/associations/deals/contacts/batch/read",
                    headers=headers,
                    json={"inputs": [{"id": did} for did in batch]},
                    timeout=15,
                )
                _vlog.warning(f"Assoc batch {batch_num}: status={assoc_resp.status_code}, attempt={attempt+1}")
                if assoc_resp.status_code in (200, 207):
                    batch_results = assoc_resp.json().get("results", [])
                    _vlog.warning(f"Assoc batch {batch_num}: got {len(batch_results)} results for {len(batch)} deals")
                    for item in batch_results:
                        did = item.get("from", {}).get("id", "")
                        for to in item.get("to", []):
                            cid = to.get("toObjectId", "")
                            if cid and did:
                                deal_contact_map.setdefault(did, []).append(str(cid))
                    break  # success, move to next batch
                elif assoc_resp.status_code == 429:
                    wait = int(assoc_resp.headers.get("Retry-After", 2)) + 1
                    _vlog.warning(f"Assoc batch {batch_num}: 429, waiting {wait}s")
                    time.sleep(wait)
                    continue  # retry this batch
                else:
                    _vlog.warning(f"Assoc batch {batch_num}: unexpected status {assoc_resp.status_code}, body={assoc_resp.text[:200]}")
                    break
            except Exception as e:
                _vlog.warning(f"Assoc batch {batch_num}: exception {e}, attempt={attempt+1}")
                if attempt < 2:
                    time.sleep(1)
                    continue
                break
        time.sleep(0.5)

    # Batch read contact properties
    all_cids = set()
    for cids in deal_contact_map.values():
        all_cids.update(cids)

    contact_props = {}
    cid_list = list(all_cids)
    for batch_start in range(0, len(cid_list), 50):
        batch = cid_list[batch_start:batch_start + 50]
        for attempt in range(3):
            try:
                cresp = req.post(
                    "https://api.hubapi.com/crm/v3/objects/contacts/batch/read",
                    headers=headers,
                    json={
                        "inputs": [{"id": cid} for cid in batch],
                        "properties": VENDOR_CONTACT_PROPERTIES,
                    },
                    timeout=15,
                )
                if cresp.status_code in (200, 207):
                    for c in cresp.json().get("results", []):
                        contact_props[c["id"]] = c.get("properties", {})
                    break
                elif cresp.status_code == 429:
                    time.sleep(int(cresp.headers.get("Retry-After", 2)) + 1)
                    continue
                else:
                    break
            except Exception:
                if attempt < 2:
                    time.sleep(1)
                    continue
                break
        time.sleep(0.5)

    # Fetch engagement notes (last activity) for deals
    deal_notes = {}
    # Skip note fetching for large result sets to avoid timeout
    if len(deals) <= 500:
        for batch_start in range(0, len(deal_ids), 25):
            batch = deal_ids[batch_start:batch_start + 25]
            try:
                note_resp = req.post(
                    "https://api.hubapi.com/crm/v4/associations/deals/notes/batch/read",
                    headers=headers,
                    json={"inputs": [{"id": did} for did in batch]},
                    timeout=15,
                )
                if note_resp.status_code in (200, 207):
                    for item in note_resp.json().get("results", []):
                        did = item.get("from", {}).get("id", "")
                        note_ids = [to.get("toObjectId", "") for to in item.get("to", [])]
                        if note_ids and did:
                            deal_notes[did] = note_ids[0]  # latest note
            except Exception:
                pass
            time.sleep(0.15)

    # Batch fetch actual note body text for deal notes
    note_bodies = {}  # note_id -> body text
    all_note_ids = list(set(str(nid) for nid in deal_notes.values() if nid))
    for batch_start in range(0, len(all_note_ids), 50):
        batch = all_note_ids[batch_start:batch_start + 50]
        try:
            nresp = req.post(
                "https://api.hubapi.com/crm/v3/objects/notes/batch/read",
                headers=headers,
                json={
                    "inputs": [{"id": nid} for nid in batch],
                    "properties": ["hs_note_body", "hs_timestamp"],
                },
                timeout=15,
            )
            if nresp.status_code in (200, 207):
                for n in nresp.json().get("results", []):
                    body = n.get("properties", {}).get("hs_note_body", "") or ""
                    # Strip HTML tags from note body
                    import re as _re
                    body = _re.sub(r"<[^>]+>", "", body).strip()
                    if body:
                        note_bodies[n["id"]] = body
            elif nresp.status_code == 429:
                time.sleep(int(nresp.headers.get("Retry-After", 2)) + 1)
        except Exception:
            pass
        time.sleep(0.15)

    # Build output deals
    output_deals = []
    by_source = {}
    by_stage = {}

    for d in deals:
        props = d.get("properties", {})
        dealname = props.get("dealname", "")
        stage_id = props.get("dealstage", "")
        stage_label = VENDOR_DEAL_STAGE_MAP.get(stage_id, stage_id)

        # Parse name from deal: "Name / Case Type - Firm"
        intake_name = dealname.split("/")[0].strip() if "/" in dealname else dealname
        case_type = ""
        parsed_source = ""
        if "/" in dealname:
            rest = dealname.split("/", 1)[1].strip()
            if " - " in rest:
                parts = rest.split(" - ", 1)
                case_type = parts[0].strip()
                parsed_source = parts[1].strip() if len(parts) > 1 else ""
            else:
                case_type = rest

        # Use HubSpot deal marketing_source property if available,
        # then fall back to contact marketing_source, then deal name parsing
        deal_ms = (props.get("marketing_source") or "").strip()
        marketing_source = deal_ms or parsed_source

        # Get contact info
        cids = deal_contact_map.get(d["id"], [])
        phone = ""
        email = ""
        special = ""
        contact_ms = ""
        if cids and cids[0] in contact_props:
            cp = contact_props[cids[0]]
            phone = cp.get("phone", "") or ""
            email = cp.get("email", "") or ""
            special = cp.get("special_circumstances", "") or ""
            contact_ms = (cp.get("marketing_source") or "").strip()

        # If deal-level source is still empty, use contact-level source
        if not marketing_source and contact_ms:
            marketing_source = contact_ms

        # Get deal engagement note body if available
        deal_note_id = str(deal_notes.get(d["id"], "")) if deal_notes.get(d["id"]) else ""
        deal_note_text = note_bodies.get(deal_note_id, "") if deal_note_id else ""

        output_deals.append({
            "id": d["id"],
            "intake_name": intake_name,
            "phone": phone,
            "contact_email": email,
            "case_type": case_type,
            "marketing_source": marketing_source,
            "source": marketing_source,  # alias for compatibility
            "create_date": props.get("createdate", ""),
            "stage": stage_label,
            "special_circumstances": special,
            "close_date": props.get("closedate", ""),
            "contact_note": deal_note_text,
            "intake_note": "",
        })

        # Stats
        src = marketing_source or "Unknown"
        by_source[src] = by_source.get(src, 0) + 1
        by_stage[stage_label] = by_stage.get(stage_label, 0) + 1

    stats = {
        "total": len(output_deals),
        "by_source": by_source,
        "by_stage": by_stage,
    }

    # Fetch contacts WITHOUT deals for this firm (by marketing_source)
    dealless_contacts = []
    try:
        existing_cids = set(all_cids)  # contacts already associated with deals
        contact_filter_groups = []
        for firm_name in firm_names:
            contact_filter_groups.append({
                "filters": [
                    {"propertyName": "marketing_source", "operator": "CONTAINS_TOKEN", "value": firm_name.split()[-1]},
                    {"propertyName": "createdate", "operator": "GTE", "value": str(start_ms)},
                    {"propertyName": "createdate", "operator": "LT", "value": str(end_ms)},
                ]
            })
        c_after = 0
        all_firm_contacts = []
        for _ in range(5):  # max 500 contacts
            cbody = {
                "filterGroups": contact_filter_groups,
                "properties": VENDOR_CONTACT_PROPERTIES + ["createdate", "lifecyclestage"],
                "sorts": [{"propertyName": "createdate", "direction": "DESCENDING"}],
                "limit": 100,
            }
            if c_after:
                cbody["after"] = str(c_after)
            try:
                cresp2 = req.post(
                    "https://api.hubapi.com/crm/v3/objects/contacts/search",
                    headers=headers, json=cbody, timeout=20
                )
                if cresp2.status_code == 200:
                    cdata = cresp2.json()
                    all_firm_contacts.extend(cdata.get("results", []))
                    c_after = cdata.get("paging", {}).get("next", {}).get("after")
                    if not c_after:
                        break
                elif cresp2.status_code == 429:
                    time.sleep(int(cresp2.headers.get("Retry-After", 2)) + 1)
                else:
                    break
            except Exception:
                break
            time.sleep(0.15)

        # Filter to only contacts NOT already in deals
        for c in all_firm_contacts:
            if c["id"] not in existing_cids:
                cp = c.get("properties", {})
                fname = cp.get("firstname", "") or ""
                lname = cp.get("lastname", "") or ""
                name = f"{fname} {lname}".strip() or "Unknown"
                dealless_contacts.append({
                    "id": c["id"],
                    "intake_name": name,
                    "phone": cp.get("phone", "") or "",
                    "contact_email": cp.get("email", "") or "",
                    "case_type": cp.get("case_type", "") or "",
                    "marketing_source": cp.get("marketing_source", "") or "",
                    "create_date": cp.get("createdate", ""),
                    "stage": cp.get("lifecyclestage", "") or "",
                    "special_circumstances": cp.get("special_circumstances", "") or "",
                    "close_date": "",
                    "contact_note": "",
                    "intake_note": "",
                })
    except Exception:
        pass

    return output_deals, stats, month_label, dealless_contacts


@app.route("/dashboard/<token>")
def vendor_dashboard_page(token):
    """Vendor-style dashboard page (Shamsi layout) for any firm token."""
    tokens = load_sales_snapshot_tokens()
    if token not in tokens:
        abort(404)

    token_data = tokens[token]
    firm_names = token_data.get("firm_names", [])
    display_name = firm_names[0] if len(firm_names) == 1 else ", ".join(firm_names)

    return render_template("vendor_dashboard.html",
                           firm_name=display_name,
                           token=token)


@app.route("/api/vendor/<token>")
def vendor_dashboard_api(token):
    """API endpoint for vendor dashboard — returns deals in Shamsi format."""
    tokens = load_sales_snapshot_tokens()
    if token not in tokens:
        return jsonify({"error": "Invalid token"}), 404

    token_data = tokens[token]
    firm_names = token_data.get("firm_names", [])
    month = int(request.args.get("month", 0))

    try:
        deals, stats, month_label, dealless_contacts = hubspot_get_vendor_deals(firm_names, month_offset=month)
        return jsonify({
            "deals": deals,
            "dealless_contacts": dealless_contacts,
            "stats": stats,
            "month_label": month_label,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as e:
        return jsonify({"error": str(e), "deals": [], "stats": {"total": 0, "by_source": {}, "by_stage": {}}}), 500


# ── Wommster Dashboard (KJ Wommster + KJ Wommster SP) ──
WOMMSTER_VIEW_TOKEN = "wm4tK9vB3hNc6jFs"

@app.route("/wommster")
def wommster_dashboard():
    token = request.args.get("token", "")
    if token and token != WOMMSTER_VIEW_TOKEN:
        abort(403)
    return render_template("vendor_dashboard.html",
                           vendor_name="Wommster",
                           subtitle="KJ Wommster + KJ Wommster SP Only",
                           api_url="/api/wommster")


@app.route("/api/wommster")
def api_wommster():
    token = request.args.get("token", "")
    if token and token != WOMMSTER_VIEW_TOKEN:
        return jsonify({"error": "Invalid token"}), 403
    try:
        month = int(request.args.get("month", 0))
        deals, stats, month_label, _ = hubspot_get_vendor_deals(
            ["KJ Injury"],
            month_offset=month,
        )
        # Filter to only Wommster sources
        wommster_sources = {"kj injury law wommster", "kj injury law wommster sp"}
        filtered = [d for d in deals if d.get("source", "").lower() in wommster_sources]
        # Recalculate stats
        by_source = {}
        by_stage = {}
        for d in filtered:
            src = d.get("source", "Unknown")
            by_source[src] = by_source.get(src, 0) + 1
            stg = d.get("stage", "Unknown")
            by_stage[stg] = by_stage.get(stg, 0) + 1
        stats = {"total": len(filtered), "by_source": by_source, "by_stage": by_stage}
        return jsonify({
            "deals": filtered,
            "stats": stats,
            "month_label": month_label,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as e:
        return jsonify({"error": str(e), "deals": [], "stats": {"total": 0, "by_source": {}, "by_stage": {}}}), 500


# ── JRE Injury Law Dashboard ──
JRE_VIEW_TOKEN = "jr3xN7kW9pLm2vQz"

@app.route("/jre")
def jre_dashboard():
    token = request.args.get("token", "")
    if token and token != JRE_VIEW_TOKEN:
        abort(403)
    return render_template("vendor_dashboard.html",
                           vendor_name="JRE Injury Law",
                           subtitle="All Sources",
                           api_url="/api/jre")


def _jre_all_time_sources():
    """Fetch all unique marketing sources for JRE deals across all time."""
    import requests as req
    if not HUBSPOT_API_KEY:
        return []
    headers = {"Authorization": f"Bearer {HUBSPOT_API_KEY}", "Content-Type": "application/json"}
    sources = set()
    after = 0
    for _ in range(10):
        body = {
            "filterGroups": [{"filters": [
                {"propertyName": "dealname", "operator": "CONTAINS_TOKEN", "value": "JRE"}
            ]}],
            "properties": ["dealname", "marketing_source"],
            "limit": 100,
        }
        if after:
            body["after"] = str(after)
        try:
            resp = req.post("https://api.hubapi.com/crm/v3/objects/deals/search",
                            headers=headers, json=body, timeout=20)
            if resp.status_code != 200:
                break
            data = resp.json()
            for d in data.get("results", []):
                props = d.get("properties", {})
                src = props.get("marketing_source") or ""
                if src:
                    sources.add(src)
                else:
                    name = props.get("dealname") or ""
                    if " - " in name:
                        sources.add(name.rsplit(" - ", 1)[-1].strip())
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after:
                break
        except Exception:
            break
    return sorted(sources) if sources else []


@app.route("/api/jre")
def api_jre():
    token = request.args.get("token", "")
    if token and token != JRE_VIEW_TOKEN:
        return jsonify({"error": "Invalid token"}), 403
    try:
        month = int(request.args.get("month", 0))
        source_filter = request.args.get("source", "").strip()
        deals, stats, month_label, _dealless = hubspot_get_vendor_deals(["JRE Injury Law"], month_offset=month)

        # Get ALL-TIME sources so the dropdown always shows every source
        all_sources = _jre_all_time_sources()
        # Also include any sources from current month deals not caught by the all-time query
        current_sources = set(d.get("marketing_source") or d.get("source") or "Unknown" for d in deals)
        merged = sorted(set(all_sources) | current_sources)
        if not merged:
            merged = ["Unknown"]

        # Apply server-side source filter if requested
        if source_filter:
            deals = [d for d in deals if (d.get("marketing_source") or d.get("source") or "Unknown") == source_filter]
            by_source = {}
            by_stage = {}
            for d in deals:
                src = d.get("marketing_source") or d.get("source") or "Unknown"
                by_source[src] = by_source.get(src, 0) + 1
                stg = d.get("stage") or "Unknown"
                by_stage[stg] = by_stage.get(stg, 0) + 1
            stats = {"total": len(deals), "by_source": by_source, "by_stage": by_stage}

        return jsonify({
            "deals": deals,
            "stats": stats,
            "sources": merged,
            "month_label": month_label,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as e:
        return jsonify({"error": str(e), "deals": [], "stats": {"total": 0, "by_source": {}, "by_stage": {}}, "sources": []}), 500


# ── Init ──
if not SHARE_TOKENS_FILE.exists():
    save_share_tokens({})

# ── Main ──
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8090))
    print(f"[Billing Portal] Starting on port {port}...")
    print(f"[Billing Portal] Data file: {DATA_FILE}")
    print(f"[Billing Portal] Share tokens: {SHARE_TOKENS_FILE}")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
